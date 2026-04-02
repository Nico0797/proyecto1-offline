from io import BytesIO
from datetime import date
from datetime import datetime

import openpyxl
import pytest

from backend.main import create_app
from backend.config import TestingConfig
from backend.database import db
from backend.models import (
    Business,
    Customer,
    Debt,
    Expense,
    LedgerEntry,
    Payment,
    Product,
    ProductMovement,
    RecurringExpense,
    Sale,
    TreasuryAccount,
)


@pytest.fixture
def app():
    app = create_app(TestingConfig)
    with app.app_context():
        db.drop_all()
        db.create_all()
        yield app
        db.session.remove()
        db.drop_all()


@pytest.fixture
def client(app):
    return app.test_client()


@pytest.fixture
def auth_token(client):
    response = client.post(
        "/api/auth/register",
        json={
            "email": "excel-validation@example.com",
            "password": "Password1!",
            "name": "Excel Validation",
        },
    )
    assert response.status_code == 201, response.get_data(as_text=True)
    return response.get_json()["access_token"]


@pytest.fixture
def business_context(client, auth_token, app):
    response = client.post(
        "/api/businesses",
        headers={"Authorization": f"Bearer {auth_token}"},
        json={
            "name": "Negocio Validación Excel",
            "currency": "COP",
            "settings": {
                "debt_term_days": 30,
                "receivables_due_soon_days": 5,
                "receivable_terms_by_sale": {},
            },
        },
    )
    assert response.status_code == 201, response.get_data(as_text=True)
    business_id = response.get_json()["business"]["id"]
    with app.app_context():
        business = Business.query.get(business_id)
        owner_user_id = business.user_id
        treasury_cash = TreasuryAccount(
            business_id=business_id,
            name="Caja principal",
            account_type="cash",
            payment_method_key="cash",
            currency=business.currency,
        )
        treasury_transfer = TreasuryAccount(
            business_id=business_id,
            name="Banco principal",
            account_type="bank",
            payment_method_key="transfer",
            currency=business.currency,
        )
        db.session.add_all([treasury_cash, treasury_transfer])
        db.session.flush()

        product_a = Product(
            business_id=business_id,
            name="Prod A",
            sku="A-001",
            type="product",
            price=1000,
            cost=400,
            stock=12,
            low_stock_threshold=5,
        )
        product_b = Product(
            business_id=business_id,
            name="Prod B",
            sku="B-001",
            type="product",
            price=2000,
            cost=900,
            stock=4,
            low_stock_threshold=5,
        )
        product_c = Product(
            business_id=business_id,
            name="Prod C",
            sku="C-001",
            type="service",
            price=800,
            cost=300,
            stock=0,
            low_stock_threshold=1,
        )
        db.session.add_all([product_a, product_b, product_c])
        db.session.flush()

        customer_named = Customer(
            business_id=business_id,
            name="Cliente Uno",
            created_by_user_id=owner_user_id,
            created_by_name="Excel Validation",
            created_by_role="PROPIETARIO",
        )
        customer_blank = Customer(
            business_id=business_id,
            name="",
            created_by_user_id=owner_user_id,
            created_by_name="Excel Validation",
            created_by_role="PROPIETARIO",
        )
        db.session.add_all([customer_named, customer_blank])
        db.session.flush()

        sale_paid = Sale(
            business_id=business_id,
            user_id=owner_user_id,
            customer_id=customer_named.id,
            sale_date=date(2026, 3, 5),
            items=[{"product_id": product_a.id, "name": "Prod A", "qty": 1, "unit_price": 1000, "total": 1000}],
            subtotal=1000,
            discount=0,
            total=1000,
            balance=0,
            collected_amount=1000,
            total_cost=400,
            payment_method="cash",
            treasury_account_id=treasury_cash.id,
            paid=True,
            note="Venta pagada",
            created_by_name="Vendedor 1",
            created_by_role="PROPIETARIO",
        )
        sale_partial = Sale(
            business_id=business_id,
            user_id=owner_user_id,
            customer_id=customer_named.id,
            sale_date=date(2026, 3, 10),
            items=[{"product_id": product_b.id, "name": "Prod B", "qty": 1, "unit_price": 2000, "total": 2000}],
            subtotal=2000,
            discount=0,
            total=2000,
            balance=500,
            collected_amount=1500,
            total_cost=900,
            payment_method="credit",
            treasury_account_id=treasury_transfer.id,
            paid=False,
            note="Venta parcial",
            created_by_name="Vendedor 2",
            created_by_role="PROPIETARIO",
        )
        sale_overdue = Sale(
            business_id=business_id,
            user_id=owner_user_id,
            customer_id=customer_blank.id,
            sale_date=date(2026, 1, 15),
            items=[{"product_id": product_c.id, "name": "Prod C", "qty": 1, "unit_price": 800, "total": 800}],
            subtotal=800,
            discount=0,
            total=800,
            balance=800,
            collected_amount=0,
            total_cost=300,
            payment_method="credit",
            paid=False,
            note="Venta vencida",
            created_by_name=None,
            created_by_role="PROPIETARIO",
        )
        sale_out_of_range = Sale(
            business_id=business_id,
            user_id=owner_user_id,
            customer_id=customer_named.id,
            sale_date=date(2026, 2, 20),
            items=[{"name": "Prod D", "qty": 1, "unit_price": 5000, "total": 5000}],
            subtotal=5000,
            discount=0,
            total=5000,
            balance=0,
            collected_amount=5000,
            total_cost=2500,
            payment_method="transfer",
            paid=True,
            note="Venta fuera de rango",
            created_by_name="Vendedor 3",
            created_by_role="PROPIETARIO",
        )
        db.session.add_all([sale_paid, sale_partial, sale_overdue, sale_out_of_range])
        db.session.flush()

        payment_partial = Payment(
            business_id=business_id,
            customer_id=customer_named.id,
            sale_id=sale_partial.id,
            payment_date=date(2026, 3, 12),
            amount=700,
            method="transfer",
            treasury_account_id=treasury_transfer.id,
            note="Abono venta parcial",
            created_by_user_id=owner_user_id,
            created_by_name="Caja 1",
            created_by_role="PROPIETARIO",
        )
        payment_general = Payment(
            business_id=business_id,
            customer_id=customer_blank.id,
            sale_id=None,
            payment_date=date(2026, 3, 15),
            amount=300,
            method="cash",
            treasury_account_id=treasury_cash.id,
            note=None,
            created_by_user_id=owner_user_id,
            created_by_name=None,
            created_by_role="PROPIETARIO",
        )
        db.session.add_all([payment_partial, payment_general])

        recurring_payable = RecurringExpense(
            business_id=business_id,
            name="Arriendo",
            amount=250,
            due_day=15,
            frequency="monthly",
            category="Servicios",
            payment_flow="payable",
            creditor_name="Arrendador",
            is_active=True,
        )
        db.session.add(recurring_payable)

        debt_operational = Debt(
            business_id=business_id,
            name="Proveedor harina",
            creditor_name="Proveedor Norte",
            category="Proveedores",
            total_amount=600,
            balance_due=400,
            due_date=date(2026, 3, 25),
            status="partial",
        )
        debt_financial = Debt(
            business_id=business_id,
            name="Préstamo equipo",
            creditor_name="Banco Centro",
            category="Prestamos",
            total_amount=1000,
            balance_due=700,
            due_date=date(2026, 4, 2),
            status="pending",
        )
        db.session.add_all([debt_operational, debt_financial])
        db.session.flush()

        expense_operational = Expense(
            business_id=business_id,
            expense_date=date(2026, 3, 8),
            category="Servicios",
            amount=300,
            description="Pago energía",
            source_type="manual",
            payment_method="cash",
            treasury_account_id=treasury_cash.id,
            created_by_user_id=owner_user_id,
            created_by_name="Admin",
            created_by_role="PROPIETARIO",
        )
        expense_supplier = Expense(
            business_id=business_id,
            expense_date=date(2026, 3, 8),
            category="Inventario",
            amount=200,
            description="Pago proveedor",
            source_type="supplier_payment",
            payment_method="transfer",
            treasury_account_id=treasury_transfer.id,
            created_by_user_id=owner_user_id,
            created_by_name="Admin",
            created_by_role="PROPIETARIO",
        )
        expense_operational_debt = Expense(
            business_id=business_id,
            expense_date=date(2026, 3, 11),
            category="Proveedores",
            amount=150,
            description="Cuota proveedor",
            source_type="debt_payment",
            payment_method="transfer",
            treasury_account_id=treasury_transfer.id,
            debt_id=debt_operational.id,
            created_by_user_id=owner_user_id,
            created_by_name="Admin",
            created_by_role="PROPIETARIO",
        )
        expense_financial_debt = Expense(
            business_id=business_id,
            expense_date=date(2026, 3, 13),
            category="Prestamos",
            amount=100,
            description="Cuota préstamo",
            source_type="debt_payment",
            payment_method="transfer",
            treasury_account_id=treasury_transfer.id,
            debt_id=debt_financial.id,
            created_by_user_id=owner_user_id,
            created_by_name="Admin",
            created_by_role="PROPIETARIO",
        )
        expense_out_of_range = Expense(
            business_id=business_id,
            expense_date=date(2026, 2, 10),
            category="Publicidad",
            amount=999,
            description="Fuera de rango",
            created_by_user_id=owner_user_id,
            created_by_name="Admin",
            created_by_role="PROPIETARIO",
        )
        db.session.add_all([
            expense_operational,
            expense_supplier,
            expense_operational_debt,
            expense_financial_debt,
            expense_out_of_range,
        ])

        movement_in = ProductMovement(
            product_id=product_a.id,
            business_id=business_id,
            user_id=owner_user_id,
            type="in",
            quantity=10,
            reason="Ajuste inicial",
            created_by_name="Admin",
            created_by_role="PROPIETARIO",
        )
        movement_sale = ProductMovement(
            product_id=product_b.id,
            business_id=business_id,
            user_id=owner_user_id,
            type="sale",
            quantity=1,
            reason="Venta mostrador",
            created_by_name="Caja 1",
            created_by_role="PROPIETARIO",
        )
        db.session.add_all([movement_in, movement_sale])

        ledger_initial_payment = LedgerEntry(
            business_id=business_id,
            customer_id=customer_named.id,
            entry_type="payment",
            amount=1000,
            entry_date=date(2026, 3, 5),
            note="Pago inicial venta",
            ref_type="sale",
            ref_id=sale_paid.id,
        )
        db.session.add(ledger_initial_payment)

        second_business = Business(user_id=business.user_id, name="Negocio Secundario", currency="COP")
        db.session.add(second_business)
        db.session.commit()

        return {
            "business_id": business_id,
            "token": auth_token,
            "headers": {"Authorization": f"Bearer {auth_token}"},
        }


@pytest.fixture
def pro_business_context(client, business_context):
    upgrade = client.post("/api/upgrade-to-pro", headers=business_context["headers"])
    assert upgrade.status_code == 200, upgrade.get_data(as_text=True)
    return business_context


def fetch_workbook(client, url, headers, query_string):
    response = client.get(url, headers=headers, query_string=query_string)
    assert response.status_code == 200, response.get_data(as_text=True)
    assert response.headers["Content-Type"].startswith("application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
    assert "attachment;" in response.headers.get("Content-Disposition", "")
    workbook = openpyxl.load_workbook(BytesIO(response.data))
    return response, workbook


def worksheet_values(ws):
    return [[cell for cell in row] for row in ws.iter_rows(values_only=True)]


def merged_ranges(ws):
    return {str(item) for item in ws.merged_cells.ranges}


def workbook_contains_marker(workbook, marker="MOTOR NUEVO OPENPYXL ACTIVO"):
    for sheet_name in workbook.sheetnames:
        ws = workbook[sheet_name]
        for row in worksheet_values(ws):
            for cell in row:
                if marker in str(cell or ""):
                    return True
    return False


def find_row_by_first_value(ws, value):
    for row in worksheet_values(ws):
        if row and row[0] == value:
            return row
    raise AssertionError(f"No row found with first value {value!r}")


def find_row_by_cell_value(ws, value):
    for row in worksheet_values(ws):
        if row and any(cell == value for cell in row):
            return row
    raise AssertionError(f"No row found containing {value!r}")


def test_combined_exports_require_pro_plan(client, business_context):
    business_id = business_context["business_id"]
    response = client.get(
        f"/api/businesses/{business_id}/export/combined",
        headers=business_context["headers"],
        query_string={"type": "finance_full", "start_date": "2026-03-01", "end_date": "2026-03-31", "direct": "1"},
    )
    assert response.status_code == 403
    payload = response.get_json()
    assert payload["upgrade_url"] == "/upgrade"


@pytest.mark.parametrize(
    ("label", "url_builder", "params", "expected_sheets", "expected_prefix"),
    [
        (
            "ventas",
            lambda business_id: f"/api/businesses/{business_id}/export/sales",
            {"start_date": "2026-03-01", "end_date": "2026-03-31", "direct": "1"},
            ["Resumen", "Detalle"],
            "VENTAS_",
        ),
        (
            "gastos",
            lambda business_id: f"/api/businesses/{business_id}/export/expenses",
            {"startDate": "2026-03-01", "endDate": "2026-03-31", "direct": "1"},
            ["Resumen", "Detalle", "Soporte"],
            "GASTOS_",
        ),
        (
            "cobros",
            lambda business_id: f"/api/businesses/{business_id}/export/combined",
            {"type": "payments_full", "start_date": "2026-03-01", "end_date": "2026-03-31", "direct": "1"},
            ["Resumen", "Detalle", "Soporte"],
            "COBROS_",
        ),
        (
            "caja",
            lambda business_id: f"/api/businesses/{business_id}/export/combined",
            {"type": "finance_full", "startDate": "2026-03-01", "endDate": "2026-03-31", "direct": "1"},
            ["Resumen", "Detalle financiero", "Cartera", "Entradas y salidas"],
            "FINANCIERO_",
        ),
    ],
)
def test_excel_exports_generate_valid_workbooks(client, business_context, label, url_builder, params, expected_sheets, expected_prefix):
    context = business_context
    if label in {"cobros", "caja"}:
        upgrade = client.post("/api/upgrade-to-pro", headers=business_context["headers"])
        assert upgrade.status_code == 200, upgrade.get_data(as_text=True)
    business_id = context["business_id"]
    response, workbook = fetch_workbook(client, url_builder(business_id), context["headers"], params)

    disposition = response.headers.get("Content-Disposition", "")
    assert expected_prefix in disposition
    assert workbook.sheetnames == expected_sheets

    for sheet_name in workbook.sheetnames:
        assert len(sheet_name) <= 31
        ws = workbook[sheet_name]
        assert ws.freeze_panes is not None
        assert ws.auto_filter.ref
        assert ws.max_row >= 1
        assert all(dim.width is None or dim.width > 0 for dim in ws.column_dimensions.values())

    summary = workbook[expected_sheets[0]]
    assert "A1:F1" in merged_ranges(summary)
    assert summary["A1"].value is not None
    assert summary["A2"].value is not None


def test_sales_report_exactness_and_filters(client, business_context):
    business_id = business_context["business_id"]
    _, workbook = fetch_workbook(
        client,
        f"/api/businesses/{business_id}/export/sales",
        business_context["headers"],
        {"start_date": "2026-03-01", "end_date": "2026-03-31", "direct": "1"},
    )
    summary = workbook["Resumen"]
    detail = workbook["Detalle"]

    assert summary["A10"].value == 3000
    assert summary["C10"].value == 2500
    assert summary["E10"].value == 500
    assert summary["A12"].value == 1500
    assert summary["C12"].value == 2

    detail_rows = worksheet_values(detail)
    detail_values = [row for row in detail_rows if row and isinstance(row[1], str) and row[1].startswith("Venta #")]
    references = {row[1] for row in detail_values}
    assert references == {"Venta #1", "Venta #2"}
    assert detail.freeze_panes == "A14"
    assert detail.auto_filter.ref.startswith("A13:")


def test_expenses_report_exactness_and_classification(client, business_context):
    business_id = business_context["business_id"]
    _, workbook = fetch_workbook(
        client,
        f"/api/businesses/{business_id}/export/expenses",
        business_context["headers"],
        {"startDate": "2026-03-01", "endDate": "2026-03-31", "direct": "1"},
    )
    summary = workbook["Resumen"]
    detail = workbook["Detalle"]
    support = workbook["Soporte"]

    assert summary["A10"].value == 750
    assert summary["C10"].value == 300
    assert summary["E10"].value == 650

    detail_rows = worksheet_values(detail)
    detail_values = [row for row in detail_rows if row and row[1] in {"Servicios", "Inventario", "Proveedores", "Prestamos"}]
    assert len(detail_values) == 4
    statuses = {row[7] for row in detail_values}
    assert "Pagado" in statuses or "Registrado" in statuses
    support_rows = worksheet_values(support)
    categories = {row[0]: row[2] for row in support_rows if row and row[0] in {"Servicios"}}
    assert categories["Servicios"] == 300


def test_payments_report_exactness_and_receivables_snapshot(client, pro_business_context):
    business_id = pro_business_context["business_id"]
    _, workbook = fetch_workbook(
        client,
        f"/api/businesses/{business_id}/export/combined",
        pro_business_context["headers"],
        {"type": "payments_full", "start_date": "2026-03-01", "end_date": "2026-03-31", "direct": "1"},
    )
    summary = workbook["Resumen"]
    detail = workbook["Detalle"]
    support = workbook["Soporte"]

    assert summary["A10"].value == "Cobros del período"
    assert summary["A11"].value == 1000
    assert summary["C11"].value == 2
    assert summary["E11"].value == 800
    assert summary["A13"].value == 0

    detail_rows = worksheet_values(detail)
    detail_values = [row for row in detail_rows if row and isinstance(row[2], str) and row[2].startswith("Cobro #")]
    assert len(detail_values) == 2
    support_rows = worksheet_values(support)
    receivable_rows = [row for row in support_rows if row and isinstance(row[1], str) and row[1].startswith("Venta #")]
    assert len(receivable_rows) == 2
    statuses = {row[5] for row in receivable_rows}
    assert statuses == {"Al día", "Vencido"}


def test_cash_report_exactness_and_no_mixed_concepts(client, pro_business_context):
    business_id = pro_business_context["business_id"]
    _, workbook = fetch_workbook(
        client,
        f"/api/businesses/{business_id}/export/combined",
        pro_business_context["headers"],
        {"type": "finance_full", "startDate": "2026-03-01", "endDate": "2026-03-31", "direct": "1"},
    )
    summary = workbook["Resumen"]
    financial = workbook["Detalle financiero"]
    receivables = workbook["Cartera"]
    movements = workbook["Entradas y salidas"]

    assert summary["A10"].value == 2000
    assert summary["C10"].value == 750
    assert summary["E10"].value == 1250
    assert summary["A12"].value == 1300
    assert summary["C12"].value == 1100

    financial_rows = worksheet_values(financial)
    financial_map = {row[0]: row[1] for row in financial_rows if row and isinstance(row[0], str) and row[0]}
    assert financial_map["Gasto operativo ejecutado"] == 300
    assert financial_map["Pagos a proveedores"] == 200
    assert financial_map["Pagos de obligaciones operativas"] == 150
    assert financial_map["Pagos de deuda financiera"] == 100
    assert financial_map["Por pagar operativo"] == 400
    assert financial_map["Deuda financiera pendiente"] == 700

    receivable_rows = worksheet_values(receivables)
    balances = [row[4] for row in receivable_rows if row and isinstance(row[1], str) and row[1].startswith("Venta #")]
    assert sorted(balances) == [500, 800]

    movement_rows = worksheet_values(movements)
    movement_values = [row for row in movement_rows if row and row[2] in {"Entrada", "Salida"}]
    income_total = sum(row[4] for row in movement_values if row[2] == "Entrada")
    expense_total = sum(row[4] for row in movement_values if row[2] == "Salida")
    assert income_total == 2000
    assert expense_total == 750
    assert all(not (row[0] is None and row[4] is None) for row in movement_values)


def test_general_business_report_exactness_filters_and_support(client, pro_business_context):
    business_id = pro_business_context["business_id"]
    _, workbook = fetch_workbook(
        client,
        f"/api/businesses/{business_id}/export/combined",
        pro_business_context["headers"],
        {"type": "general_business", "startDate": "2026-03-01", "endDate": "2026-03-31", "direct": "1"},
    )
    summary = workbook["Resumen"]
    detail = workbook["Detalle"]
    support = workbook["Soporte"]

    assert summary["A10"].value == 3000
    assert summary["C10"].value == 2500
    assert summary["E10"].value == 750
    assert summary["A12"].value == 1000
    assert summary["C12"].value == 1300
    assert find_row_by_first_value(summary, "Ventas registradas")[1] == 2
    assert find_row_by_first_value(summary, "Ticket promedio")[1] == 1500
    assert find_row_by_first_value(summary, "Clientes con saldo")[1] == 2
    assert find_row_by_first_value(summary, "Resultado bruto ventas-gastos")[1] == 2250
    assert summary.freeze_panes == "A16"
    assert summary.auto_filter.ref.startswith("A15:")
    assert find_row_by_cell_value(summary, "2026-03-01")
    assert find_row_by_cell_value(summary, "2026-03-31")

    detail_rows = worksheet_values(detail)
    day_rows = {row[0].date(): row for row in detail_rows if row and isinstance(row[0], datetime)}
    assert day_rows[date(2026, 3, 10)][1:] == [2000, 1, 1500, 500, 0, 2000]
    assert day_rows[date(2026, 3, 8)][1:] == [0, 0, 0, 0, 500, -500]
    assert detail.freeze_panes == "A14"
    assert detail.auto_filter.ref.startswith("A13:")

    support_rows = worksheet_values(support)
    cliente_uno_row = next(row for row in support_rows if row and row[0] == "Cliente" and row[1] == "Cliente Uno")
    prod_b_row = next(row for row in support_rows if row and row[0] == "Producto" and row[1] == "Prod B")
    assert cliente_uno_row[2:5] == [3000, 2, 500]
    assert prod_b_row[2:5] == [2000, 1, 1]


def test_customers_360_report_exactness_and_cutoff_consistency(client, pro_business_context):
    business_id = pro_business_context["business_id"]
    _, workbook = fetch_workbook(
        client,
        f"/api/businesses/{business_id}/export/combined",
        pro_business_context["headers"],
        {"type": "customers_full", "startDate": "2026-03-01", "endDate": "2026-03-31", "direct": "1"},
    )
    summary = workbook["Resumen"]
    detail = workbook["Detalle"]
    support = workbook["Soporte"]

    assert find_row_by_first_value(summary, "Clientes con ventas")[1] == 1
    assert find_row_by_first_value(summary, "Ventas acumuladas")[1] == 3000
    assert find_row_by_first_value(summary, "Cobros acumulados")[1] == 1000
    assert find_row_by_cell_value(summary, 800)
    assert find_row_by_cell_value(summary, "Corte cartera")
    assert find_row_by_cell_value(summary, "2026-03-31")

    detail_rows = worksheet_values(detail)
    cliente_uno = next(row for row in detail_rows if row and row[0] == "Cliente Uno")
    assert cliente_uno[2:7] == ["Al día", 2, 3000, 700, 500]
    assert detail.freeze_panes == "A14"
    assert detail.auto_filter.ref.startswith("A13:")

    support_rows = worksheet_values(support)
    casual_support = next(row for row in support_rows if row and row[0] == "Cliente casual")
    assert casual_support[1:] == ["Venta #3", datetime(2026, 1, 15, 0, 0), datetime(2026, 2, 14, 0, 0), 800, "Vencido"]
    refs = {row[1]: row[5] for row in support_rows if row and isinstance(row[1], str) and row[1].startswith("Venta #")}
    assert refs == {"Venta #2": "Al día", "Venta #3": "Vencido"}


def test_inventory_report_exactness_and_rotation_support(client, pro_business_context):
    business_id = pro_business_context["business_id"]
    _, workbook = fetch_workbook(
        client,
        f"/api/businesses/{business_id}/export/combined",
        pro_business_context["headers"],
        {"type": "products_full", "startDate": "2026-03-01", "endDate": "2026-03-31", "direct": "1"},
    )
    summary = workbook["Resumen"]
    detail = workbook["Detalle"]
    support = workbook["Soporte"]

    assert summary["A10"].value == 3
    assert summary["C10"].value == 16
    assert summary["E10"].value == 8400
    assert summary["A12"].value == 20000
    assert find_row_by_first_value(summary, "Productos con stock bajo")[1] == 2
    assert find_row_by_first_value(summary, "Costo promedio por producto")[1] == 2800
    assert round(float(find_row_by_first_value(summary, "Precio promedio por producto")[1]), 2) == 1266.67

    detail_rows = worksheet_values(detail)
    statuses = {row[0]: row[7] for row in detail_rows if row and row[0] in {"Prod A", "Prod B", "Prod C"}}
    assert statuses == {"Prod A": "OK", "Prod B": "Bajo", "Prod C": "Agotado"}
    prod_b = next(row for row in detail_rows if row and row[0] == "Prod B")
    assert prod_b[3:10] == [2000, 900, 4, 5, "Bajo", 3600, 8000]

    support_rows = worksheet_values(support)
    movement_products = {row[1] for row in support_rows if row and row[1] in {"Prod A", "Prod B"}}
    assert movement_products == {"Prod A", "Prod B"}
    assert support.freeze_panes == "A14"
    assert support.auto_filter.ref.startswith("A13:")


def test_aged_receivables_exactness_and_bucket_consistency(client, pro_business_context):
    business_id = pro_business_context["business_id"]
    _, workbook = fetch_workbook(
        client,
        f"/api/businesses/{business_id}/export/combined",
        pro_business_context["headers"],
        {"type": "aged_receivables", "startDate": "2026-03-01", "endDate": "2026-03-31", "direct": "1"},
    )
    summary = workbook["Resumen"]
    detail = workbook["Detalle"]
    support = workbook["Soporte"]

    assert find_row_by_first_value(summary, "Clientes con saldo")[1] == 2
    assert find_row_by_first_value(summary, "Ventas abiertas")[1] == 2
    assert find_row_by_first_value(summary, "Cartera vencida")[1] == 800
    assert find_row_by_cell_value(summary, 1300)
    assert find_row_by_cell_value(summary, 500)

    detail_rows = worksheet_values(detail)
    casual = next(row for row in detail_rows if row and row[0] == "Cliente casual")
    uno = next(row for row in detail_rows if row and row[0] == "Cliente Uno")
    assert casual[1:] == [0, 0, 800, 0, 0, 800]
    assert uno[1:] == [500, 0, 0, 0, 0, 500]

    support_rows = worksheet_values(support)
    buckets = {row[1]: row[6] for row in support_rows if row and isinstance(row[1], str) and row[1].startswith("Venta #")}
    assert buckets == {"Venta #2": "Al día", "Venta #3": "31-60 días"}


def test_profitability_export_exactness_and_query_filter_consistency(client, pro_business_context):
    business_id = pro_business_context["business_id"]
    _, workbook = fetch_workbook(
        client,
        f"/api/businesses/{business_id}/export/profitability",
        pro_business_context["headers"],
        {
            "start_date": "2026-03-01",
            "end_date": "2026-03-31",
            "status": "all",
            "product_query": "Prod A",
            "focus": "top_products",
            "direct": "1",
        },
    )
    summary = workbook["Resumen"]
    products = workbook["Productos"]
    sales = workbook["Ventas"]

    assert find_row_by_cell_value(summary, 3000)
    assert find_row_by_first_value(summary, "Costo consumido")[1] == 0
    assert find_row_by_first_value(summary, "Ventas sin consumo")[1] == 2
    assert find_row_by_cell_value(summary, "Prod A")
    assert find_row_by_cell_value(summary, "top_products")

    product_rows = [row for row in worksheet_values(products) if row and row[0] == "Prod A"]
    assert len(product_rows) == 1
    assert product_rows[0][1] == "Sin consumo relacionado"
    assert product_rows[0][4:8] == [1000, 0, 0, 0]

    sales_rows = [row for row in worksheet_values(sales) if row and isinstance(row[0], str) and row[0].startswith("Venta #")]
    assert sales_rows == []
    assert find_row_by_cell_value(sales, "No hay ventas de rentabilidad para el rango seleccionado.")


def test_team_export_exactness_and_visible_filters(client, pro_business_context):
    business_id = pro_business_context["business_id"]
    _, workbook = fetch_workbook(
        client,
        f"/api/businesses/{business_id}/export/team",
        pro_business_context["headers"],
        {"start_date": "2026-03-01", "end_date": "2026-03-31", "direct": "1"},
    )
    summary = workbook["Resumen"]
    detail = workbook["Detalle"]

    assert find_row_by_cell_value(summary, 1)
    assert find_row_by_cell_value(summary, 3000)
    assert find_row_by_cell_value(summary, 1000)
    assert find_row_by_cell_value(summary, "2026-03-01")
    assert find_row_by_cell_value(summary, "2026-03-31")

    summary_rows = worksheet_values(summary)
    owner_row = next(row for row in summary_rows if row and row[1] == "Vendedor 1")
    assert owner_row[3:12] == [2, 3000, 2, 1000, 4, 750, 2, 2, 0]

    detail_rows = [row for row in worksheet_values(detail) if row and isinstance(row[3], str)]
    actions = {row[3] for row in detail_rows}
    assert "Venta" in actions
    assert "Recaudo" in actions
    assert "Gasto" in actions
    assert "Cliente Nuevo" in actions
    assert any(action.startswith("Inventario") for action in actions)
    assert detail.freeze_panes == "A14"
    assert detail.auto_filter.ref.startswith("A13:")


@pytest.mark.parametrize(
    ("report_type", "expected_sheets", "expected_prefix"),
    [
        ("general_business", ["Resumen", "Detalle", "Soporte"], "GENERAL_"),
        ("customers_full", ["Resumen", "Detalle", "Soporte"], "CLIENTES_CARTERA_"),
        ("products_full", ["Resumen", "Detalle", "Soporte"], "INVENTARIO_"),
        ("profit_full", ["Resumen", "Productos", "Ventas", "Alertas", "Soporte"], "RENTABILIDAD_"),
        ("aged_receivables", ["Resumen", "Detalle", "Soporte"], "CARTERA_EDADES_"),
        ("cashflow_full", ["Resumen", "Detalle financiero", "Cartera", "Entradas y salidas"], "FLUJO_CAJA_"),
    ],
)
def test_current_catalog_combined_exports_use_new_engine_marker(client, pro_business_context, report_type, expected_sheets, expected_prefix):
    business_id = pro_business_context["business_id"]
    response, workbook = fetch_workbook(
        client,
        f"/api/businesses/{business_id}/export/combined",
        pro_business_context["headers"],
        {"type": report_type, "start_date": "2026-03-01", "end_date": "2026-03-31", "direct": "1"},
    )
    assert expected_prefix in response.headers.get("Content-Disposition", "")
    assert workbook.sheetnames == expected_sheets
    assert workbook_contains_marker(workbook)


def test_profitability_direct_export_uses_new_engine_marker(client, pro_business_context):
    business_id = pro_business_context["business_id"]
    response, workbook = fetch_workbook(
        client,
        f"/api/businesses/{business_id}/export/profitability",
        pro_business_context["headers"],
        {"start_date": "2026-03-01", "end_date": "2026-03-31", "direct": "1"},
    )
    assert "RENTABILIDAD_" in response.headers.get("Content-Disposition", "")
    assert workbook.sheetnames == ["Resumen", "Productos", "Ventas", "Alertas", "Soporte"]
    assert workbook_contains_marker(workbook)


def test_team_export_uses_new_engine_marker(client, pro_business_context):
    business_id = pro_business_context["business_id"]
    response, workbook = fetch_workbook(
        client,
        f"/api/businesses/{business_id}/export/team",
        pro_business_context["headers"],
        {"start_date": "2026-03-01", "end_date": "2026-03-31", "direct": "1"},
    )
    assert "REPORTE_EQUIPO_" in response.headers.get("Content-Disposition", "")
    assert workbook.sheetnames == ["Resumen", "Detalle"]
    assert workbook_contains_marker(workbook)


def test_empty_and_out_of_range_exports_handle_gracefully(client, pro_business_context):
    business_id = pro_business_context["business_id"]
    cases = [
        (f"/api/businesses/{business_id}/export/sales", {"start_date": "2027-01-01", "end_date": "2027-01-31", "direct": "1"}, "Detalle"),
        (f"/api/businesses/{business_id}/export/expenses", {"start_date": "2027-01-01", "end_date": "2027-01-31", "direct": "1"}, "Detalle"),
        (f"/api/businesses/{business_id}/export/combined", {"type": "payments_full", "start_date": "2027-01-01", "end_date": "2027-01-31", "direct": "1"}, "Detalle"),
        (f"/api/businesses/{business_id}/export/combined", {"type": "finance_full", "start_date": "2027-01-01", "end_date": "2027-01-31", "direct": "1"}, "Entradas y salidas"),
    ]
    for url, params, sheet_name in cases:
        _, workbook = fetch_workbook(client, url, pro_business_context["headers"], params)
        ws = workbook[sheet_name]
        values = [row for row in worksheet_values(ws) if row and any(cell is not None for cell in row)]
        assert any("No hay" in str(cell) for row in values for cell in row if cell is not None)


def test_exports_are_isolated_by_business(client, auth_token, business_context, app):
    with app.app_context():
        second_business = Business.query.filter_by(name="Negocio Secundario").first()
        extra_sale = Sale(
            business_id=second_business.id,
            sale_date=date(2026, 3, 9),
            items=[{"name": "Otro", "qty": 1, "unit_price": 9999, "total": 9999}],
            subtotal=9999,
            discount=0,
            total=9999,
            balance=0,
            collected_amount=9999,
            total_cost=0,
            payment_method="cash",
            paid=True,
        )
        db.session.add(extra_sale)
        db.session.commit()

    business_id = business_context["business_id"]
    _, workbook = fetch_workbook(
        client,
        f"/api/businesses/{business_id}/export/sales",
        business_context["headers"],
        {"start_date": "2026-03-01", "end_date": "2026-03-31", "direct": "1"},
    )
    detail_rows = worksheet_values(workbook["Detalle"])
    refs = {row[1] for row in detail_rows if row and isinstance(row[1], str) and row[1].startswith("Venta #")}
    assert "Venta #5" not in refs
