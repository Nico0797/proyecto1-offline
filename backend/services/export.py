import os
import json
from datetime import datetime, date
from sqlalchemy import func, case, desc
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
from flask import current_app

from backend.database import db
from backend.models import (
    Business, Product, Customer, Sale, Expense, 
    Payment, LedgerEntry, ProductMovement, User, Debt, DebtPayment
)
from backend.services.analytics_layer import AnalyticsLayer

class ExcelReportGenerator:
    def __init__(self, business_id, user_id=None):
        self.business_id = business_id
        self.user_id = user_id
        self.wb = Workbook()
        self.analytics = AnalyticsLayer(business_id)
        
        # Eliminar hoja por defecto
        if "Sheet" in self.wb.sheetnames:
            del self.wb["Sheet"]
            
        # Estilos Estándar
        self.header_font = Font(bold=True, color="FFFFFF", size=11)
        self.header_fill = PatternFill(start_color="1F2937", end_color="1F2937", fill_type="solid") # Dark Gray
        self.currency_fmt = '"$"#,##0.00_-'
        self.date_fmt = 'yyyy-mm-dd'
        self.border = Border(
            left=Side(style='thin', color='D1D5DB'),
            right=Side(style='thin', color='D1D5DB'),
            top=Side(style='thin', color='D1D5DB'),
            bottom=Side(style='thin', color='D1D5DB')
        )

    def _create_sheet(self, title, headers, tab_color=None):
        ws = self.wb.create_sheet(title=title[:31]) # Excel limit 31 chars
        if tab_color:
            ws.sheet_properties.tabColor = tab_color
            
        # Escribir encabezados
        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_num, value=header)
            cell.font = self.header_font
            cell.fill = self.header_fill
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border = self.border
            
        # Congelar fila superior
        ws.freeze_panes = "A2"
        # Añadir filtro automático
        ws.auto_filter.ref = ws.dimensions
        
        return ws

    def _auto_adjust_columns(self, ws):
        for column in ws.columns:
            max_length = 0
            column_letter = get_column_letter(column[0].column)
            for cell in column:
                try:
                    if cell.value:
                        val_len = len(str(cell.value))
                        if val_len > max_length:
                            max_length = val_len
                except:
                    pass
            adjusted_width = min(max_length + 4, 60)
            ws.column_dimensions[column_letter].width = adjusted_width

    def save(self, prefix):
        export_dir = current_app.config.get("EXPORT_DIR", "exports")
        if not os.path.isabs(export_dir):
            export_dir = os.path.join(current_app.root_path, export_dir)
        os.makedirs(export_dir, exist_ok=True)
        
        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
        filename = f"{prefix}_{self.business_id}_{timestamp}.xlsx"
        filepath = os.path.join(export_dir, filename)
        
        self.wb.save(filepath)
        return filepath

    def _get_debt_scope(self, category):
        normalized = str(category or "").strip().lower()
        return "financial" if normalized in {"tarjetas", "prestamos", "financiaciones", "creditos", "leasing"} else "operational"

    def _classify_financial_expenses(self, expenses):
        debt_ids = sorted({expense.debt_id for expense in expenses if expense.debt_id})
        debts_by_id = {}
        if debt_ids:
            debts_by_id = {
                debt.id: debt
                for debt in db.session.query(Debt).filter(Debt.business_id == self.business_id, Debt.id.in_(debt_ids)).all()
            }

        totals = {
            "operational_expense": 0.0,
            "supplier_payment": 0.0,
            "operational_obligation_payment": 0.0,
            "financial_debt_payment": 0.0,
        }
        operational_categories = {}
        detail_rows = []

        for expense in expenses:
            amount = float(expense.amount or 0)
            source_type = str(expense.source_type or "manual").strip().lower() or "manual"
            scope = ""
            flow_label = "Gasto operativo ejecutado"

            if source_type == "supplier_payment":
                totals["supplier_payment"] += amount
                scope = "operational"
                flow_label = "Pago a proveedor"
            elif source_type == "debt_payment":
                related_debt = debts_by_id.get(expense.debt_id)
                scope = self._get_debt_scope(related_debt.category if related_debt else None)
                if scope == "financial":
                    totals["financial_debt_payment"] += amount
                    flow_label = "Pago de deuda financiera"
                else:
                    totals["operational_obligation_payment"] += amount
                    flow_label = "Pago de obligación operativa"
            else:
                totals["operational_expense"] += amount
                scope = "operational"
                if source_type == "purchase_payment":
                    flow_label = "Compra pagada"
                elif source_type == "recurring":
                    flow_label = "Gasto recurrente ejecutado"
                category_label = expense.category or "Sin categoría"
                bucket = operational_categories.setdefault(category_label, {"total": 0.0, "count": 0})
                bucket["total"] += amount
                bucket["count"] += 1

            detail_rows.append({
                "date": expense.expense_date,
                "flow_label": flow_label,
                "scope": "Financiero" if scope == "financial" else "Operativo",
                "category": expense.category or "Sin categoría",
                "description": expense.description or flow_label,
                "amount": amount,
                "source_type": source_type,
            })

        total_real_out = sum(totals.values())
        return {
            "totals": {key: round(value, 2) for key, value in totals.items()},
            "total_real_out": round(total_real_out, 2),
            "operational_categories": operational_categories,
            "detail_rows": detail_rows,
        }

    # --- REPORTES ESPECÍFICOS ---

    def generate_general_business(self, start_date=None, end_date=None):
        """1. REPORTE GENERAL DEL NEGOCIO"""
        
        # Consultas Base
        sales_q = db.session.query(Sale).filter(Sale.business_id == self.business_id)
        expenses_q = db.session.query(Expense).filter(Expense.business_id == self.business_id)
        
        if start_date:
            sales_q = sales_q.filter(Sale.sale_date >= start_date)
            expenses_q = expenses_q.filter(Expense.expense_date >= start_date)
        if end_date:
            sales_q = sales_q.filter(Sale.sale_date <= end_date)
            expenses_q = expenses_q.filter(Expense.expense_date <= end_date)

        # KPIs
        total_sales = sales_q.with_entities(func.sum(Sale.total)).scalar() or 0
        total_expenses = expenses_q.with_entities(func.sum(Expense.amount)).scalar() or 0
        net_income = total_sales - total_expenses
        margin = (net_income / total_sales * 100) if total_sales > 0 else 0
        tx_count = sales_q.count()
        avg_ticket = total_sales / tx_count if tx_count > 0 else 0
        
        # Hoja 1: Resumen Ejecutivo
        ws = self._create_sheet("Resumen Ejecutivo", ["Indicador", "Valor", "Notas"], "1E40AF")
        
        kpis = [
            ("Ventas Totales", total_sales, "Ingresos brutos"),
            ("Gastos Operativos", total_expenses, "Egresos registrados"),
            ("Utilidad Neta", net_income, "Ganancia real"),
            ("Margen de Utilidad", f"{margin:.2f}%", "Rentabilidad sobre ventas"),
            ("Transacciones", tx_count, "Cantidad de ventas"),
            ("Ticket Promedio", avg_ticket, "Venta promedio por cliente")
        ]
        
        for row_idx, (label, val, note) in enumerate(kpis, 2):
            ws.cell(row=row_idx, column=1, value=label)
            cell = ws.cell(row=row_idx, column=2, value=val)
            if isinstance(val, (int, float)) and "%" not in str(val):
                cell.number_format = self.currency_fmt
            ws.cell(row=row_idx, column=3, value=note)

        self._auto_adjust_columns(ws)
        
        # Hoja 2: Ventas por Método de Pago
        ws_pay = self._create_sheet("Métodos de Pago", ["Método", "Total", "Transacciones"], "2563EB")
        methods = sales_q.with_entities(
            Sale.payment_method, 
            func.sum(Sale.total),
            func.count(Sale.id)
        ).group_by(Sale.payment_method).all()
        
        for m_name, m_total, m_count in methods:
            ws_pay.append([m_name or "Desconocido", m_total, m_count])
            ws_pay.cell(row=ws_pay.max_row, column=2).number_format = self.currency_fmt
            
        self._auto_adjust_columns(ws_pay)

        # Hoja 3: Top Productos (General)
        ws_prod = self._create_sheet("Top Productos", ["Producto", "Cant. Vendida", "Total Ventas"], "D97706")
        top_prods = db.session.query(
            Sale.items
        ).filter(Sale.business_id == self.business_id)
        
        if start_date: top_prods = top_prods.filter(Sale.sale_date >= start_date)
        if end_date: top_prods = top_prods.filter(Sale.sale_date <= end_date)
        
        # Procesamiento en Python para items JSON (limitación de SQLite/JSON)
        prod_stats = {}
        for s in top_prods.all():
            if s.items and isinstance(s.items, list):
                for item in s.items:
                    if isinstance(item, dict):
                        name = item.get("name", "Unknown")
                        qty = float(item.get("qty", 0))
                        total = float(item.get("total", 0))
                        if name not in prod_stats: prod_stats[name] = {"qty": 0, "total": 0}
                        prod_stats[name]["qty"] += qty
                        prod_stats[name]["total"] += total
        
        sorted_prods = sorted(prod_stats.items(), key=lambda x: x[1]["total"], reverse=True)[:20]
        for name, stats in sorted_prods:
            ws_prod.append([name, stats["qty"], stats["total"]])
            ws_prod.cell(row=ws_prod.max_row, column=3).number_format = self.currency_fmt
        self._auto_adjust_columns(ws_prod)

    def generate_finance_report(self, start_date=None, end_date=None):
        expenses_q = db.session.query(Expense).filter(Expense.business_id == self.business_id)
        if start_date:
            expenses_q = expenses_q.filter(Expense.expense_date >= start_date)
        if end_date:
            expenses_q = expenses_q.filter(Expense.expense_date <= end_date)

        expenses = expenses_q.order_by(Expense.expense_date.desc(), Expense.id.desc()).all()
        expense_breakdown = self._classify_financial_expenses(expenses)

        payments_q = db.session.query(Payment).filter(Payment.business_id == self.business_id)
        cash_sales_q = db.session.query(Sale).filter(Sale.business_id == self.business_id, Sale.payment_method == "cash")
        debts_q = db.session.query(Debt).filter(Debt.business_id == self.business_id)
        if start_date:
            payments_q = payments_q.filter(Payment.payment_date >= start_date)
            cash_sales_q = cash_sales_q.filter(Sale.sale_date >= start_date)
        if end_date:
            payments_q = payments_q.filter(Payment.payment_date <= end_date)
            cash_sales_q = cash_sales_q.filter(Sale.sale_date <= end_date)

        cash_in_total = float(payments_q.with_entities(func.sum(Payment.amount)).scalar() or 0) + float(cash_sales_q.with_entities(func.sum(Sale.total)).scalar() or 0)
        cash_out_total = float(expense_breakdown["total_real_out"] or 0)
        cash_net_total = cash_in_total - cash_out_total

        debts = debts_q.all()
        operational_debts = [debt for debt in debts if debt.status != "paid" and self._get_debt_scope(debt.category) == "operational"]
        financial_debts = [debt for debt in debts if debt.status != "paid" and self._get_debt_scope(debt.category) == "financial"]

        ws_summary = self._create_sheet("Resumen Financiero", ["Indicador", "Valor", "Lectura"], "15803D")
        summary_rows = [
            ("Entradas reales ejecutadas", cash_in_total, "Cobros registrados en el período"),
            ("Salidas reales ejecutadas", cash_out_total, "No incluye pendientes; solo movimientos ya ejecutados"),
            ("Flujo neto real", cash_net_total, "Entradas reales menos salidas reales"),
            ("Gasto operativo ejecutado", expense_breakdown["totals"]["operational_expense"], "Gasto manual, recurrente ejecutado y compras pagadas"),
            ("Pagos a proveedores", expense_breakdown["totals"]["supplier_payment"], "Cancelaciones reales de cuentas por pagar a proveedor"),
            ("Pagos de obligaciones operativas", expense_breakdown["totals"]["operational_obligation_payment"], "Pagos reales de obligaciones operativas no financieras"),
            ("Pagos de deuda financiera", expense_breakdown["totals"]["financial_debt_payment"], "Pagos reales de tarjetas, préstamos y créditos"),
            ("Por pagar operativo pendiente", sum(float(debt.balance_due or 0) for debt in operational_debts), "Obligaciones operativas aún no ejecutadas"),
            ("Deuda financiera pendiente", sum(float(debt.balance_due or 0) for debt in financial_debts), "Pasivos financieros aún no cancelados"),
        ]
        for label, value, note in summary_rows:
            ws_summary.append([label, value, note])
            ws_summary.cell(row=ws_summary.max_row, column=2).number_format = self.currency_fmt
        self._auto_adjust_columns(ws_summary)

        ws_breakdown = self._create_sheet("Salidas Reales", ["Tipo de salida", "Monto Total", "% de salidas reales"], "B91C1C")
        breakdown_rows = [
            ("Gasto operativo ejecutado", expense_breakdown["totals"]["operational_expense"]),
            ("Pagos a proveedores", expense_breakdown["totals"]["supplier_payment"]),
            ("Pagos de obligaciones operativas", expense_breakdown["totals"]["operational_obligation_payment"]),
            ("Pagos de deuda financiera", expense_breakdown["totals"]["financial_debt_payment"]),
        ]
        for label, amount in breakdown_rows:
            pct = (amount / cash_out_total) if cash_out_total > 0 else 0
            ws_breakdown.append([label, amount, pct])
            ws_breakdown.cell(row=ws_breakdown.max_row, column=2).number_format = self.currency_fmt
            ws_breakdown.cell(row=ws_breakdown.max_row, column=3).number_format = '0.00%'
        self._auto_adjust_columns(ws_breakdown)

        ws_detail = self._create_sheet("Detalle Flujos", ["Fecha", "Tipo de flujo", "Alcance", "Categoría", "Descripción", "Monto", "Origen"], "1D4ED8")
        for row in expense_breakdown["detail_rows"]:
            ws_detail.append([
                row["date"],
                row["flow_label"],
                row["scope"],
                row["category"],
                row["description"],
                row["amount"],
                row["source_type"],
            ])
            ws_detail.cell(row=ws_detail.max_row, column=6).number_format = self.currency_fmt
        self._auto_adjust_columns(ws_detail)

        ws_operational = self._create_sheet("Gasto Operativo", ["Categoría", "Monto Total", "% del gasto operativo", "Registros"], "DC2626")
        operational_total = float(expense_breakdown["totals"]["operational_expense"] or 0)
        for category, data in sorted(expense_breakdown["operational_categories"].items(), key=lambda item: item[1]["total"], reverse=True):
            pct = (data["total"] / operational_total) if operational_total > 0 else 0
            ws_operational.append([category, data["total"], pct, data["count"]])
            ws_operational.cell(row=ws_operational.max_row, column=2).number_format = self.currency_fmt
            ws_operational.cell(row=ws_operational.max_row, column=3).number_format = '0.00%'
        self._auto_adjust_columns(ws_operational)

        ws_operational_payables = self._create_sheet("Por Pagar Operativo", ["Nombre", "Acreedor", "Categoría", "Vencimiento", "Estado", "Saldo"], "F59E0B")
        for debt in sorted(operational_debts, key=lambda item: (item.due_date or date.max, item.id)):
            ws_operational_payables.append([
                debt.name,
                debt.creditor_name or "-",
                debt.category or "Sin categoría",
                debt.due_date,
                debt.status,
                debt.balance_due,
            ])
            ws_operational_payables.cell(row=ws_operational_payables.max_row, column=6).number_format = self.currency_fmt
        self._auto_adjust_columns(ws_operational_payables)

        ws_financial_debt = self._create_sheet("Deuda Financiera", ["Nombre", "Acreedor", "Categoría", "Vencimiento", "Estado", "Saldo"], "7C3AED")
        for debt in sorted(financial_debts, key=lambda item: (item.due_date or date.max, item.id)):
            ws_financial_debt.append([
                debt.name,
                debt.creditor_name or "-",
                debt.category or "Sin categoría",
                debt.due_date,
                debt.status,
                debt.balance_due,
            ])
            ws_financial_debt.cell(row=ws_financial_debt.max_row, column=6).number_format = self.currency_fmt
        self._auto_adjust_columns(ws_financial_debt)

    def generate_sales_report(self, start_date=None, end_date=None):
        """3. REPORTE DE VENTAS DETALLADO"""
        
        sales_q = db.session.query(Sale).filter(Sale.business_id == self.business_id)
        if start_date: sales_q = sales_q.filter(Sale.sale_date >= start_date)
        if end_date: sales_q = sales_q.filter(Sale.sale_date <= end_date)
        
        sales = sales_q.order_by(Sale.sale_date.desc()).all()
        
        # Hoja 1: Detalle Transaccional
        headers = ["ID", "Fecha", "Cliente", "Estado Pago", "Método", "Subtotal", "Descuento", "Total", "Productos (Resumen)", "Notas", "Vendedor", "Rol"]
        ws = self._create_sheet("Ventas Detalladas", headers, "059669")
        
        for s in sales:
            c_name = s.customer.name if s.customer else "Cliente General"
            status = "PAGADO" if s.paid else "PENDIENTE"
            
            # Resumen de productos
            items_str = ""
            if s.items and isinstance(s.items, list):
                items_str = ", ".join([f"{i.get('qty')}x {i.get('name')}" for i in s.items if isinstance(i, dict)])
            
            ws.append([
                s.id, s.sale_date, c_name, status, s.payment_method, 
                s.subtotal, s.discount, s.total, items_str, s.note,
                s.created_by_name or "Histórico", s.created_by_role or "-"
            ])
            # Formatos
            row = ws.max_row
            for col in [6, 7, 8]: # Money cols
                ws.cell(row=row, column=col).number_format = self.currency_fmt
                
        self._auto_adjust_columns(ws)

    def generate_inventory_report(self):
        """4. REPORTE DE INVENTARIO Y MOVIMIENTOS"""
        
        products = Product.query.filter_by(business_id=self.business_id).all()
        
        # Hoja 1: Estado Actual
        headers = ["Producto", "SKU", "Categoría", "Costo", "Precio", "Stock Actual", "Alerta Stock", "Valor Inventario (Costo)", "Valor Venta Potencial"]
        ws = self._create_sheet("Valorización Inventario", headers, "D97706")
        
        for p in products:
            val_cost = (p.cost or 0) * p.stock
            val_price = p.price * p.stock
            stock_alert = "BAJO" if p.stock <= p.low_stock_threshold else "OK"
            if p.stock <= 0: stock_alert = "AGOTADO"
            
            ws.append([
                p.name, p.sku, "General", p.cost, p.price, p.stock, 
                stock_alert, val_cost, val_price
            ])
            
            row = ws.max_row
            for col in [4, 5, 8, 9]:
                ws.cell(row=row, column=col).number_format = self.currency_fmt
                
        self._auto_adjust_columns(ws)
        
        # Hoja 2: Movimientos Recientes (Últimos 1000)
        movements = db.session.query(ProductMovement).filter(
            ProductMovement.business_id == self.business_id
        ).order_by(ProductMovement.created_at.desc()).limit(1000).all()
        
        ws_mov = self._create_sheet("Kardex (Últimos 1000)", ["Fecha", "Producto", "Tipo", "Cantidad", "Motivo", "Usuario", "Rol"], "F59E0B")
        
        for m in movements:
            p_name = m.product.name if m.product else "Eliminado"
            user_display = m.created_by_name if m.created_by_name else (m.user.name if m.user else "Sistema")
            role_display = m.created_by_role if m.created_by_role else "-"
            
            ws_mov.append([
                m.created_at, p_name, m.type, m.quantity, m.reason, user_display, role_display
            ])
            
        self._auto_adjust_columns(ws_mov)

    def generate_receivables_report(self):
        """5. REPORTE DE CARTERA / CUENTAS POR COBRAR"""
        
        customers = Customer.query.filter_by(business_id=self.business_id).all()
        
        headers = ["Cliente", "Teléfono", "Saldo Total", "Última Compra", "Último Pago", "Estado"]
        ws = self._create_sheet("Cartera de Clientes", headers, "7C3AED")
        
        for c in customers:
            charges = db.session.query(func.sum(LedgerEntry.amount)).filter(
                LedgerEntry.customer_id == c.id, LedgerEntry.entry_type == 'charge'
            ).scalar() or 0
            
            payments = db.session.query(func.sum(LedgerEntry.amount)).filter(
                LedgerEntry.customer_id == c.id, LedgerEntry.entry_type == 'payment'
            ).scalar() or 0
            
            balance = charges - payments
            
            if abs(balance) > 100:
                last_sale = Sale.query.filter_by(customer_id=c.id).order_by(Sale.sale_date.desc()).first()
                last_payment = Payment.query.filter_by(customer_id=c.id).order_by(Payment.payment_date.desc()).first()
                
                status = "AL DÍA"
                if balance > 0: status = "DEUDA"
                if balance < 0: status = "A FAVOR"
                
                ws.append([
                    c.name, c.phone, balance, 
                    last_sale.sale_date if last_sale else None,
                    last_payment.payment_date if last_payment else None,
                    status
                ])
                ws.cell(row=ws.max_row, column=3).number_format = self.currency_fmt
                
        self._auto_adjust_columns(ws)

    # --- NUEVOS REPORTES AVANZADOS ---

    def generate_profitability_report(self, start_date=None, end_date=None):
        """6. REPORTE DE RENTABILIDAD POR PRODUCTO"""
        
        # Obtener datos de AnalyticsLayer
        # TODO: Mover la lógica de analytics_layer a get_profitability_data y llamarla aquí
        from backend.services.analytics_layer import AnalyticsLayer
        analytics = AnalyticsLayer(self.business_id)
        data = analytics.get_profitability_data(start_date, end_date)
        
        headers = ["Producto", "Categoría", "Cantidad Vendida", "Venta Total (Ingreso)", "Costo Total", "Utilidad Bruta", "Margen %"]
        ws = self._create_sheet("Rentabilidad por Producto", headers, "10B981")
        
        for item in data:
            margin = (item["profit"] / item["revenue"] * 100) if item["revenue"] > 0 else 0
            
            ws.append([
                item["name"],
                item["category"],
                item["qty"],
                item["revenue"],
                item["cost"],
                item["profit"],
                margin
            ])
            
            row = ws.max_row
            for col in [4, 5, 6]: # Money cols
                ws.cell(row=row, column=col).number_format = self.currency_fmt
            ws.cell(row=row, column=7).number_format = '0.00%'
            
        self._auto_adjust_columns(ws)

    def generate_aged_receivables_report(self):
        """7. CARTERA POR EDADES (VENCIMIENTO)"""
        
        from backend.services.analytics_layer import AnalyticsLayer
        analytics = AnalyticsLayer(self.business_id)
        data = analytics.get_aged_receivables()
        
        headers = ["Cliente", "Saldo Total", "Corriente (0-30 días)", "31-60 días", "61-90 días", "+90 días"]
        ws = self._create_sheet("Cartera por Edades", headers, "EF4444")
        
        for item in data:
            ws.append([
                item["customer_name"],
                item["total"],
                item["0-30"],
                item["31-60"],
                item["61-90"],
                item["+90"]
            ])
            
            row = ws.max_row
            for col in range(2, 7): # Money cols
                ws.cell(row=row, column=col).number_format = self.currency_fmt
                
        self._auto_adjust_columns(ws)

    def generate_cashflow_report(self, start_date, end_date):
        from backend.services.analytics_layer import AnalyticsLayer
        analytics = AnalyticsLayer(self.business_id)
        data = analytics.get_cashflow_data(start_date, end_date)

        headers = [
            "Fecha",
            "Entradas reales",
            "Gasto operativo ejecutado",
            "Pagos a proveedores",
            "Pagos obligaciones operativas",
            "Pagos deuda financiera",
            "Salidas reales",
            "Flujo neto diario",
        ]
        ws = self._create_sheet("Flujo Caja Real", headers, "3B82F6")

        for item in data:
            ws.append([
                item["date"],
                item["income"],
                item.get("operational_expense", 0),
                item.get("supplier_payment", 0),
                item.get("operational_obligation_payment", 0),
                item.get("financial_debt_payment", 0),
                item["outcome"],
                item["net"]
            ])
            
            row = ws.max_row
            for col in [2, 3, 4, 5, 6, 7, 8]:
                ws.cell(row=row, column=col).number_format = self.currency_fmt
                
        self._auto_adjust_columns(ws)

    def generate_team_report(self, summary_data, detail_data, start_date=None, end_date=None):
        wb = Workbook()
        
        # Hoja 1: Resumen
        ws = wb.active
        ws.title = "Resumen Equipo"
        
        headers = ["ID", "Nombre", "Rol", "Ventas (#)", "Ventas ($)", "Recaudos (#)", "Recaudos ($)", "Gastos (#)", "Gastos ($)", "Clientes Nuevos", "Mov. Inv.", "Recordatorios"]
        ws.append(headers)
        
        # Header Style
        fill = PatternFill(start_color="1E3A8A", end_color="1E3A8A", fill_type="solid")
        font = Font(color="FFFFFF", bold=True)
        for cell in ws[1]:
            cell.fill = fill
            cell.font = font
            
        for s in summary_data:
            ws.append([
                s["user_id"],
                s["name"],
                s["role"],
                s["sales_count"],
                s["sales_total"],
                s["payments_count"],
                s["payments_total"],
                s["expenses_count"],
                s["expenses_total"],
                s["customers_created"],
                s["movements_count"],
                s["reminders_created"]
            ])
            
            # Format Money columns (5, 7, 9) -> indices 4, 6, 8 (0-based) but excel uses 1-based
            row = ws.max_row
            ws.cell(row=row, column=5).number_format = self.currency_fmt
            ws.cell(row=row, column=7).number_format = self.currency_fmt
            ws.cell(row=row, column=9).number_format = self.currency_fmt
            
        self._auto_adjust_columns(ws)
        
        # Hoja 2: Detalle Actividad
        ws_det = wb.create_sheet("Detalle Actividad")
        ws_det.append(["Fecha", "Empleado", "Rol", "Acción", "Referencia", "Valor ($)", "Detalle"])
        for cell in ws_det[1]:
            cell.fill = fill
            cell.font = font
            
        for d in detail_data:
            ws_det.append([
                d["date"],
                d["user_name"],
                d["user_role"],
                d["action"],
                d["reference"],
                d["amount"],
                d["detail"]
            ])
            # Formato moneda col 6
            ws_det.cell(row=ws_det.max_row, column=6).number_format = self.currency_fmt
            
        self._auto_adjust_columns(ws_det)
            
        return wb

    def generate_profitability_report(self, summary_data, products_data, sales_data, alerts_data, start_date=None, end_date=None, filters=None):
        summary_ws = self._create_sheet("Resumen Rentabilidad", ["Indicador", "Valor", "Contexto"], "059669")

        period_label = f"{start_date or 'Inicio'} a {end_date or 'Fin'}"
        summary_rows = [
            ("Periodo", period_label, "Rango exportado"),
            ("Ventas Totales", summary_data.get("revenue_total"), "Ventas registradas en el período"),
            ("Ventas Costeadas", summary_data.get("costed_revenue_total"), "Ventas con margen confiable"),
            ("Costo Consumido", summary_data.get("consumed_cost_total"), "Costo real consumido en ventas completas"),
            ("Margen Bruto Estimado", summary_data.get("gross_margin_total"), "Solo sobre ventas completas"),
            ("Margen % Estimado", f"{summary_data.get('margin_percent')}%" if summary_data.get("margin_percent") is not None else "No estimable", "No se inventa si faltan datos"),
            ("Ventas Completas", summary_data.get("complete_sales_count"), "Con margen confiable"),
            ("Ventas Incompletas", summary_data.get("incomplete_sales_count"), "Costo parcial o faltante"),
            ("Ventas Sin Consumo", summary_data.get("no_consumption_sales_count"), "Sin consumo relacionado"),
            ("Ventas Sin Costo Base", summary_data.get("missing_cost_sales_count"), "Sin costo base suficiente"),
            ("Productos con Advertencia", summary_data.get("products_with_issues_count"), "Incompletos, sin costo o sin consumo"),
        ]

        filters = filters or {}
        if filters.get("status"):
            summary_rows.append(("Filtro Estado", filters.get("status"), "Aplicado desde la vista"))
        if filters.get("product_query"):
            summary_rows.append(("Filtro Producto", filters.get("product_query"), "Aplicado desde la vista"))
        if filters.get("focus"):
            summary_rows.append(("Foco", filters.get("focus"), "Origen de navegación"))

        for row_idx, (label, value, note) in enumerate(summary_rows, 2):
            summary_ws.cell(row=row_idx, column=1, value=label)
            value_cell = summary_ws.cell(row=row_idx, column=2, value=value)
            if isinstance(value, (int, float)):
                value_cell.number_format = self.currency_fmt
            summary_ws.cell(row=row_idx, column=3, value=note)
        self._auto_adjust_columns(summary_ws)

        alerts_ws = self._create_sheet("Alertas", ["Nivel", "Código", "Título", "Mensaje", "Casos"], "D97706")
        for alert in alerts_data.get("alerts", []):
            alerts_ws.append([
                alert.get("level"),
                alert.get("code"),
                alert.get("title"),
                alert.get("message"),
                alert.get("count"),
            ])
        self._auto_adjust_columns(alerts_ws)

        products_ws = self._create_sheet(
            "Productos Rentabilidad",
            [
                "Producto",
                "Estado",
                "Mensaje",
                "Cantidad Vendida",
                "Ventas Totales",
                "Ventas Costeadas",
                "Costo Consumido",
                "Margen Estimado",
                "Margen %",
                "Ventas",
                "Ventas Costeadas #",
                "Incompletas #",
                "Sin Consumo #",
                "Sin Costo #",
            ],
            "2563EB",
        )
        for item in products_data.get("items", []):
            products_ws.append([
                item.get("product_name"),
                item.get("cost_status_label") or item.get("cost_status"),
                item.get("cost_status_message"),
                item.get("quantity_sold"),
                item.get("revenue_total"),
                item.get("costed_revenue_total"),
                item.get("consumed_cost_total"),
                item.get("estimated_gross_margin"),
                item.get("estimated_margin_percent"),
                item.get("sales_count"),
                item.get("costed_sales_count"),
                item.get("incomplete_sales_count"),
                item.get("no_consumption_sales_count"),
                item.get("missing_cost_sales_count"),
            ])
            row = products_ws.max_row
            for col in [5, 6, 7, 8]:
                products_ws.cell(row=row, column=col).number_format = self.currency_fmt
        self._auto_adjust_columns(products_ws)

        sales_ws = self._create_sheet(
            "Ventas Rentabilidad",
            [
                "Venta",
                "Fecha",
                "Cliente",
                "Estado",
                "Mensaje",
                "Método Pago",
                "Total Venta",
                "Costo Consumido",
                "Costo Parcial",
                "Margen Estimado",
                "Margen %",
                "Items",
                "Consumos",
            ],
            "7C3AED",
        )
        for item in sales_data.get("items", []):
            sales_ws.append([
                item.get("sale_id"),
                item.get("sale_date"),
                item.get("customer_name") or "Cliente casual",
                item.get("cost_status_label") or item.get("cost_status"),
                item.get("cost_status_message"),
                item.get("payment_method"),
                item.get("sale_total"),
                item.get("consumed_cost_total"),
                item.get("partial_consumed_cost_total"),
                item.get("estimated_gross_margin"),
                item.get("estimated_margin_percent"),
                item.get("items_count"),
                item.get("consumptions_count"),
            ])
            row = sales_ws.max_row
            for col in [7, 8, 9, 10]:
                sales_ws.cell(row=row, column=col).number_format = self.currency_fmt
        self._auto_adjust_columns(sales_ws)

        issues_ws = self._create_sheet(
            "Casos con Advertencia",
            ["Tipo", "Referencia", "Estado", "Mensaje", "Monto/Ventas", "Margen", "Fecha"],
            "DC2626",
        )
        for item in products_data.get("incomplete_items", []):
            issues_ws.append([
                "Producto",
                item.get("product_name"),
                item.get("cost_status_label") or item.get("cost_status"),
                item.get("cost_status_message"),
                item.get("revenue_total"),
                item.get("estimated_margin_percent"),
                None,
            ])
            issues_ws.cell(row=issues_ws.max_row, column=5).number_format = self.currency_fmt
        for item in sales_data.get("no_consumption_items", []) + sales_data.get("incomplete_items", []):
            issues_ws.append([
                "Venta",
                f"Venta #{item.get('sale_id')}",
                item.get("cost_status_label") or item.get("cost_status"),
                item.get("cost_status_message"),
                item.get("sale_total"),
                item.get("estimated_margin_percent"),
                item.get("sale_date"),
            ])
            issues_ws.cell(row=issues_ws.max_row, column=5).number_format = self.currency_fmt
        self._auto_adjust_columns(issues_ws)


def export_combined_report(business_id, report_type, start_date=None, end_date=None):
    """Router principal de exportación"""
    normalized_type = str(report_type or "").strip().lower()
    if normalized_type in {"sales_full", "payments_full", "finance_full", "cashflow_full"}:
        from backend.services.reports.report_service import export_combined_report as professional_export_combined_report

        return professional_export_combined_report(business_id, report_type, start_date, end_date)
    
    # Parse Fechas
    start = None
    end = None
    if start_date:
        try: start = datetime.strptime(str(start_date)[:10], "%Y-%m-%d").date()
        except: pass
    if end_date:
        try: end = datetime.strptime(str(end_date)[:10], "%Y-%m-%d").date()
        except: pass

    generator = ExcelReportGenerator(business_id)
    filename_prefix = "REPORTE"

    if report_type == "general_business":
        generator.generate_general_business(start, end)
        filename_prefix = "GENERAL"
        
    elif report_type == "finance_full":
        generator.generate_finance_report(start, end)
        filename_prefix = "FINANCIERO"
        
    elif report_type == "customers_full":
        generator.generate_receivables_report() # Incluye cartera
        filename_prefix = "CLIENTES_CARTERA"
        
    elif report_type == "products_full":
        generator.generate_inventory_report()
        filename_prefix = "INVENTARIO"

    elif report_type == "sales_full":
        generator.generate_sales_report(start, end)
        filename_prefix = "VENTAS_DETALLE"
        
    elif report_type == "profit_full":
        generator.generate_profitability_report(start, end)
        filename_prefix = "RENTABILIDAD"
        
    elif report_type == "aged_receivables":
        generator.generate_aged_receivables_report()
        filename_prefix = "CARTERA_EDADES"
        
    elif report_type == "cashflow_full":
        generator.generate_cashflow_report(start, end)
        filename_prefix = "FLUJO_CAJA"
        
    else:
        # Fallback para compatibilidad
        generator.generate_general_business(start, end)
        
    return generator.save(filename_prefix)

# Wrappers legacy para mantener compatibilidad si se llaman directamente
def export_sales_excel(business_id, start_date=None, end_date=None):
    from backend.services.reports.report_service import export_sales_excel as professional_export_sales_excel

    return professional_export_sales_excel(business_id, start_date, end_date)

def export_expenses_excel(business_id, start_date=None, end_date=None):
    from backend.services.reports.report_service import export_expenses_excel as professional_export_expenses_excel

    return professional_export_expenses_excel(business_id, start_date, end_date)

def export_payments_excel(business_id, start_date=None, end_date=None):
    from backend.services.reports.report_service import export_payments_excel as professional_export_payments_excel

    return professional_export_payments_excel(business_id, start_date, end_date)

def export_cash_excel(business_id, start_date=None, end_date=None):
    from backend.services.reports.report_service import export_cash_excel as professional_export_cash_excel

    return professional_export_cash_excel(business_id, start_date, end_date)

def export_profitability_excel(business_id, summary_data, products_data, sales_data, alerts_data, start_date=None, end_date=None, filters=None):
    from backend.services.reports.report_service import export_profitability_excel as professional_export_profitability_excel

    return professional_export_profitability_excel(
        business_id,
        summary_data,
        products_data,
        sales_data,
        alerts_data,
        start_date,
        end_date,
        filters=filters,
    )
