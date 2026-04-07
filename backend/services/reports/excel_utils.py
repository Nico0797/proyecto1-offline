from datetime import datetime
from typing import Iterable, Sequence

from openpyxl.workbook import Workbook
from openpyxl.worksheet.worksheet import Worksheet
from openpyxl.utils import get_column_letter


def safe_sheet_title(title: str) -> str:
    cleaned = str(title or "Hoja").strip() or "Hoja"
    for invalid_char in [":", "\\", "/", "?", "*", "[", "]"]:
        cleaned = cleaned.replace(invalid_char, "-")
    return cleaned[:31]


def ensure_workbook_ready(workbook: Workbook):
    if "Sheet" in workbook.sheetnames and len(workbook.sheetnames) == 1:
        del workbook["Sheet"]


def configure_workbook_properties(workbook: Workbook, title: str, subject: str, creator: str = "Cuaderno Contable"):
    workbook.properties.title = title
    workbook.properties.subject = subject
    workbook.properties.creator = creator
    workbook.properties.lastModifiedBy = creator
    workbook.properties.company = "Cuaderno Contable"
    workbook.properties.created = datetime.utcnow()
    workbook.properties.modified = datetime.utcnow()


def merge_labeled_row(ws: Worksheet, row: int, start_column: int, end_column: int, value: str, palette, style_name: str):
    ws.merge_cells(start_row=row, start_column=start_column, end_row=row, end_column=end_column)
    cell = ws.cell(row=row, column=start_column, value=value)
    palette.apply_named(cell, style_name)
    return cell


def write_filters_table(ws: Worksheet, start_row: int, filters: Sequence[tuple[str, object]], palette, max_pairs_per_row: int = 2) -> int:
    row = start_row
    pair_index = 0
    for label, value in filters:
        if pair_index and pair_index % max_pairs_per_row == 0:
            row += 1
        base_column = (pair_index % max_pairs_per_row) * 2 + 1
        label_cell = ws.cell(row=row, column=base_column, value=label)
        value_cell = ws.cell(row=row, column=base_column + 1, value=value)
        palette.apply_named(label_cell, "filter_label")
        palette.apply_named(value_cell, "filter_value")
        pair_index += 1
    return row


def write_kpi_block(ws: Worksheet, start_row: int, kpis: Sequence[dict], palette, columns_per_row: int = 3) -> int:
    if not kpis:
        return start_row

    row = start_row
    for index, item in enumerate(kpis):
        if index and index % columns_per_row == 0:
            row += 2
        block_index = index % columns_per_row
        start_col = block_index * 2 + 1
        label_cell = ws.cell(row=row, column=start_col, value=item.get("label"))
        value_cell = ws.cell(row=row + 1, column=start_col, value=item.get("value"))
        palette.apply_named(label_cell, "kpi_label")
        palette.apply_named(value_cell, "kpi_value", number_format=item.get("number_format"))
        ws.merge_cells(start_row=row, start_column=start_col, end_row=row, end_column=start_col + 1)
        ws.merge_cells(start_row=row + 1, start_column=start_col, end_row=row + 1, end_column=start_col + 1)
    return row + 1


def write_table_headers(ws: Worksheet, row: int, headers: Sequence[str], palette):
    for column, header in enumerate(headers, start=1):
        cell = ws.cell(row=row, column=column, value=header)
        palette.apply_named(cell, "table_header")
    ws.row_dimensions[row].height = 22


def write_empty_state(ws: Worksheet, row: int, message: str, total_columns: int, palette):
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=max(1, total_columns))
    cell = ws.cell(row=row, column=1, value=message)
    palette.apply_named(cell, "empty_state")
    ws.row_dimensions[row].height = 28


def write_footer(ws: Worksheet, row: int, text: str, palette, total_columns: int):
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=max(1, total_columns))
    cell = ws.cell(row=row, column=1, value=text)
    palette.apply_named(cell, "footer")


def apply_freeze_filter(ws: Worksheet, header_row: int, total_columns: int):
    ws.freeze_panes = f"A{header_row + 1}"
    ws.auto_filter.ref = f"A{header_row}:{get_column_letter(max(total_columns, 1))}{max(header_row, ws.max_row)}"


def autosize_columns(ws: Worksheet, min_width: int = 12, max_width: int = 40, extra_padding: int = 2, fixed_widths: dict[str, int] | None = None):
    fixed_widths = fixed_widths or {}
    for col_idx in range(1, ws.max_column + 1):
        letter = get_column_letter(col_idx)
        if letter in fixed_widths:
            ws.column_dimensions[letter].width = fixed_widths[letter]
            continue
        max_length = 0
        for row in range(1, ws.max_row + 1):
            value = ws.cell(row=row, column=col_idx).value
            if value is None:
                continue
            text = str(value)
            text_length = max(len(chunk) for chunk in text.splitlines()) if text else 0
            if text_length > max_length:
                max_length = text_length
        ws.column_dimensions[letter].width = max(min_width, min(max_width, max_length + extra_padding))


def configure_print(ws: Worksheet, landscape: bool = True, repeat_header_row: int | None = None):
    ws.sheet_view.showGridLines = False
    ws.page_setup.orientation = "landscape" if landscape else "portrait"
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 0
    ws.print_options.horizontalCentered = False
    ws.print_options.verticalCentered = False
    ws.page_margins.left = 0.3
    ws.page_margins.right = 0.3
    ws.page_margins.top = 0.5
    ws.page_margins.bottom = 0.5
    ws.page_margins.header = 0.2
    ws.page_margins.footer = 0.2
    if repeat_header_row:
        ws.print_title_rows = f"{repeat_header_row}:{repeat_header_row}"


def set_row_number_formats(ws: Worksheet, row: int, formats: dict[int, str]):
    for column_index, number_format in formats.items():
        ws.cell(row=row, column=column_index).number_format = number_format


def append_total_row(ws: Worksheet, row: int, values: Sequence[object], palette, number_formats: dict[int, str] | None = None):
    number_formats = number_formats or {}
    for column, value in enumerate(values, start=1):
        cell = ws.cell(row=row, column=column, value=value)
        if column == 1:
            palette.apply_named(cell, "total_label")
        else:
            palette.apply_named(cell, "total_value", number_format=number_formats.get(column))


def prepare_standard_layout(
    ws: Worksheet,
    palette,
    title: str,
    subtitle: str,
    filters: Sequence[tuple[str, object]] | None,
    kpis: Sequence[dict] | None,
    table_title: str,
    headers: Sequence[str],
):
    max_cols = max(6, len(headers))
    merge_labeled_row(ws, 1, 1, max_cols, title, palette, "title")
    merge_labeled_row(ws, 2, 1, max_cols, subtitle, palette, "subtitle")

    row = 4
    if filters:
        merge_labeled_row(ws, row, 1, max_cols, "Filtros aplicados", palette, "section_title")
        row = write_filters_table(ws, row + 1, filters, palette) + 2

    if kpis:
        merge_labeled_row(ws, row, 1, max_cols, "Resumen", palette, "section_title")
        row = write_kpi_block(ws, row + 1, kpis, palette) + 2

    merge_labeled_row(ws, row, 1, max_cols, table_title, palette, "section_title")
    header_row = row + 1
    write_table_headers(ws, header_row, headers, palette)
    return header_row


def add_generated_footer(ws: Worksheet, palette, total_columns: int, business_name: str):
    footer_row = ws.max_row + 2
    timestamp = datetime.now().strftime("%Y-%m-%d %H:%M")
    write_footer(ws, footer_row, f"Generado por Cuaderno Contable para {business_name} el {timestamp} | MOTOR NUEVO OPENPYXL ACTIVO", palette, total_columns)
