import os
from datetime import date, datetime, timedelta

from flask import current_app
from openpyxl import Workbook
from sqlalchemy import func
from sqlalchemy.orm import joinedload

from backend.database import db
from backend.models import Business, Debt, DebtPayment, Expense, LedgerEntry, Payment, Product, RecurringExpense, Sale

from .excel_styles import ExcelStylePalette
from .excel_utils import configure_workbook_properties, ensure_workbook_ready, safe_sheet_title

from .cash_report import build_cash_report
from .catalog_report_builders import (
    build_aged_receivables_report,
    build_customers_report,
    build_general_business_report,
    build_inventory_report,
    build_profitability_payload_report,
    build_team_report,
)
from .expenses_report import build_expenses_report
from .payments_report import build_payments_report
from .sales_report import build_sales_report


class ReportExportService:
    def __init__(self, business_id: int, user_id: int | None = None):
        self.business_id = business_id
        self.user_id = user_id
        self.business = Business.query.get(business_id)
        if not self.business:
            raise ValueError("Negocio no encontrado")

        self.workbook = Workbook()
        ensure_workbook_ready(self.workbook)
        self.styles = ExcelStylePalette(self.business.currency or "COP")
        configure_workbook_properties(
            self.workbook,
            title=f"Reportes {self.business.name}",
            subject=f"Exportaciones Excel de {self.business.name}",
        )

    def create_sheet(self, title: str, tab_color: str | None = None):
        ws = self.workbook.create_sheet(title=safe_sheet_title(title))
        if tab_color:
            ws.sheet_properties.tabColor = tab_color
        return ws

    def save(self, prefix: str):
        export_dir = current_app.config.get("EXPORT_DIR", "exports")
        if not os.path.isabs(export_dir):
            export_dir = os.path.join(current_app.root_path, export_dir)
        os.makedirs(export_dir, exist_ok=True)
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        filepath = os.path.join(export_dir, f"{prefix}_{self.business_id}_{timestamp}.xlsx")
        self.workbook.save(filepath)
        return filepath

    @staticmethod
    def parse_date(value):
        if value in (None, ""):
            return None
        if isinstance(value, date) and not isinstance(value, datetime):
            return value
        if isinstance(value, datetime):
            return value.date()
        try:
            return datetime.strptime(str(value)[:10], "%Y-%m-%d").date()
        except Exception:
            return None

    def normalize_dates(self, start_date=None, end_date=None):
        start = self.parse_date(start_date)
        end = self.parse_date(end_date)
        if start and end and start > end:
            raise ValueError("El rango de fechas es inválido")
        return start, end

    def build_filter_rows(self, start_date=None, end_date=None, extra_filters: dict | None = None):
        filters = [
            ("Negocio", self.business.name),
            ("Moneda", self.styles.currency_code),
            ("Fecha inicio", start_date.isoformat() if start_date else "Sin límite"),
            ("Fecha fin", end_date.isoformat() if end_date else "Sin límite"),
        ]
        for key, value in (extra_filters or {}).items():
            filters.append((key, value if value not in (None, "") else "No aplica"))
        return filters

    def period_label(self, start_date=None, end_date=None):
        return f"{start_date.isoformat() if start_date else 'Inicio'} a {end_date.isoformat() if end_date else 'Hoy'}"

    def get_debt_scope(self, category):
        normalized = str(category or "").strip().lower()
        return "financial" if normalized in {"tarjetas", "prestamos", "financiaciones", "creditos", "leasing"} else "operational"

    def build_debts_summary(self, debts, today=None):
        reference = today or date.today()
        active_debts = [debt for debt in debts if (debt.status or "pending") != "paid" and float(debt.balance_due or 0) > 0]
        overdue_debts = [debt for debt in active_debts if debt.due_date and debt.due_date < reference]
        due_today_debts = [debt for debt in active_debts if debt.due_date == reference]
        due_soon_debts = [
            debt for debt in active_debts
            if debt.due_date and debt.due_date > reference and (debt.due_date - reference).days <= 7
        ]
        month_start = date(reference.year, reference.month, 1)
        next_month_start = date(reference.year + 1, 1, 1) if reference.month == 12 else date(reference.year, reference.month + 1, 1)
        paid_this_month = 0.0
        if debts:
            paid_this_month = float(
                db.session.query(func.coalesce(func.sum(DebtPayment.amount), 0))
                .join(Debt, Debt.id == DebtPayment.debt_id)
                .filter(
                    Debt.business_id == self.business_id,
                    DebtPayment.payment_date >= month_start,
                    DebtPayment.payment_date < next_month_start,
                )
                .scalar() or 0
            )
        return {
            "total_debt": round(sum(float(item.balance_due or 0) for item in active_debts), 2),
            "active_count": len(active_debts),
            "overdue_total": round(sum(float(item.balance_due or 0) for item in overdue_debts), 2),
            "overdue_count": len(overdue_debts),
            "due_today_total": round(sum(float(item.balance_due or 0) for item in due_today_debts), 2),
            "due_soon_total": round(sum(float(item.balance_due or 0) for item in due_soon_debts), 2),
            "paid_this_month": round(paid_this_month, 2),
        }

    def _get_receivables_settings(self):
        settings = self.business.settings or {}
        default_term_days = int(settings.get("debt_term_days") or 30)
        if default_term_days < 0:
            default_term_days = 30
        due_soon_days = int(settings.get("receivables_due_soon_days") or 5)
        if due_soon_days < 1:
            due_soon_days = 5
        if default_term_days > 0:
            due_soon_days = min(due_soon_days, default_term_days)
        sale_term_overrides = settings.get("receivable_terms_by_sale") or {}
        if not isinstance(sale_term_overrides, dict):
            sale_term_overrides = {}
        return default_term_days, due_soon_days, sale_term_overrides

    def resolve_sale_due_date(self, sale):
        default_term_days, _, sale_term_overrides = self._get_receivables_settings()
        override = sale_term_overrides.get(str(sale.id))
        try:
            term_days = int(override) if override not in (None, "") else int(default_term_days)
        except Exception:
            term_days = int(default_term_days)
        if term_days < 0:
            term_days = int(default_term_days)
        if not sale.sale_date:
            return None
        return sale.sale_date + timedelta(days=term_days)

    def build_receivables_snapshot(self, reference_date=None):
        today = reference_date or date.today()
        _, due_soon_days, _ = self._get_receivables_settings()
        sales = (
            Sale.query.options(joinedload(Sale.customer))
            .filter(Sale.business_id == self.business_id, Sale.sale_date <= today, Sale.balance > 0)
            .order_by(Sale.sale_date.desc(), Sale.id.desc())
            .all()
        )

        items = []
        customer_ids = set()
        overdue_customer_ids = set()
        summary = {
            "total_pending": 0.0,
            "customers_with_balance": 0,
            "overdue_customers_count": 0,
            "open_count": 0,
            "overdue_total": 0.0,
            "due_today_total": 0.0,
            "due_soon_total": 0.0,
            "current_total": 0.0,
        }

        for sale in sales:
            balance = round(float(sale.balance or 0), 2)
            if balance <= 0:
                continue
            due_date = self.resolve_sale_due_date(sale)
            customer_name = sale.customer.name if sale.customer and sale.customer.name else "Cliente casual"
            status = "Al día"
            days_until_due = None
            if due_date:
                days_until_due = (due_date - today).days
                if due_date < today:
                    status = "Vencido"
                    summary["overdue_total"] += balance
                    if sale.customer_id:
                        overdue_customer_ids.add(sale.customer_id)
                elif due_date == today:
                    status = "Vence hoy"
                    summary["due_today_total"] += balance
                elif days_until_due <= due_soon_days:
                    status = "Por vencer"
                    summary["due_soon_total"] += balance
                else:
                    summary["current_total"] += balance
            else:
                summary["current_total"] += balance

            summary["total_pending"] += balance
            summary["open_count"] += 1
            if sale.customer_id:
                customer_ids.add(sale.customer_id)

            items.append({
                "sale_id": sale.id,
                "reference": f"Venta #{sale.id}",
                "sale_date": sale.sale_date,
                "due_date": due_date,
                "customer_name": customer_name,
                "sale_total": round(float(sale.total or 0), 2),
                "collected_amount": round(float(sale.collected_amount or 0), 2),
                "balance": balance,
                "status": status,
                "days_until_due": days_until_due,
            })

        summary["customers_with_balance"] = len(customer_ids)
        summary["overdue_customers_count"] = len(overdue_customer_ids)
        for key in list(summary.keys()):
            if key.endswith("_total"):
                summary[key] = round(float(summary[key] or 0), 2)
        return {"summary": summary, "items": items}

    def resolve_sale_cost_total(self, sale, products_map=None):
        stored_cost = round(float(sale.total_cost or 0), 2)
        if stored_cost > 0:
            return stored_cost, "sale_total_cost"

        items = sale.items or []
        if not items:
            return None, "missing"

        fallback_cost = 0.0
        for item in items:
            product_id = item.get("product_id")
            if not product_id:
                return None, "missing"
            product = products_map.get(int(product_id)) if products_map else Product.query.get(int(product_id))
            unit_cost = float(getattr(product, "cost", 0) or 0) if product else 0
            if unit_cost <= 0:
                return None, "missing"
            quantity = float(item.get("quantity") if item.get("quantity") is not None else item.get("qty") or 0)
            if quantity <= 0:
                return None, "missing"
            fallback_cost += unit_cost * quantity

        return round(fallback_cost, 2), "product_cost_fallback"

    def _payment_application_label(self, payment):
        if payment.sale_id:
            return f"Aplicado a venta #{payment.sale_id}"
        note = str(payment.note or "").strip()
        return note or "Saldo general del cliente"

    def get_sales_dataset(self, start_date=None, end_date=None):
        query = Sale.query.options(joinedload(Sale.customer)).filter(Sale.business_id == self.business_id)
        if start_date:
            query = query.filter(Sale.sale_date >= start_date)
        if end_date:
            query = query.filter(Sale.sale_date <= end_date)
        sales = query.order_by(Sale.sale_date.desc(), Sale.id.desc()).all()

        product_ids = set()
        for sale in sales:
            for item in sale.items or []:
                product_id = item.get("product_id")
                if product_id:
                    product_ids.add(int(product_id))
        products_map = {}
        if product_ids:
            products_map = {
                product.id: product
                for product in Product.query.filter(Product.id.in_(sorted(product_ids))).all()
            }

        rows = []
        total_sales = 0.0
        total_collected = 0.0
        total_pending = 0.0
        total_cost = 0.0
        costed_sales_total = 0.0
        uncosted_sales_total = 0.0
        missing_cost_sales_count = 0
        for sale in sales:
            total = round(float(sale.total or 0), 2)
            collected = round(float(sale.collected_amount or 0), 2)
            pending = round(float(sale.balance or 0), 2)
            sale_cost, cost_source = self.resolve_sale_cost_total(sale, products_map)
            total_sales += total
            total_collected += collected
            total_pending += pending
            if sale_cost is not None:
                total_cost += sale_cost
                costed_sales_total += total
            else:
                uncosted_sales_total += total
                missing_cost_sales_count += 1
            rows.append({
                "date": sale.sale_date,
                "reference": f"Venta #{sale.id}",
                "customer": sale.customer.name if sale.customer else "Cliente casual",
                "status": "Pagada" if pending <= 0.01 else ("Parcial" if collected > 0 else "Pendiente"),
                "total": total,
                "cost_total": sale_cost,
                "cost_source": cost_source,
                "collected": collected,
                "pending": pending,
                "payment_method": sale.payment_method or "-",
                "seller": sale.created_by_name or "Histórico",
                "note": sale.note or "",
            })
        count = len(rows)
        return {
            "rows": rows,
            "summary": {
                "sales_total": round(total_sales, 2),
                "total_cost": round(total_cost, 2),
                "costed_sales_total": round(costed_sales_total, 2),
                "uncosted_sales_total": round(uncosted_sales_total, 2),
                "missing_cost_sales_count": int(missing_cost_sales_count),
                "sales_collected": round(total_collected, 2),
                "sales_pending": round(total_pending, 2),
                "average_ticket": round(total_sales / count, 2) if count else 0.0,
                "sales_count": count,
            },
        }

    def get_payments_dataset(self, start_date=None, end_date=None):
        query = Payment.query.options(joinedload(Payment.customer)).filter(Payment.business_id == self.business_id)
        if start_date:
            query = query.filter(Payment.payment_date >= start_date)
        if end_date:
            query = query.filter(Payment.payment_date <= end_date)
        payments = query.order_by(Payment.payment_date.desc(), Payment.id.desc()).all()
        receivables = self.build_receivables_snapshot(reference_date=end_date or date.today())

        rows = []
        total_received = 0.0
        for payment in payments:
            amount = round(float(payment.amount or 0), 2)
            total_received += amount
            rows.append({
                "customer": payment.customer.name if payment.customer else f"Cliente #{payment.customer_id}",
                "payment_date": payment.payment_date,
                "reference": f"Cobro #{payment.id}",
                "method": payment.method or "-",
                "amount": amount,
                "application": self._payment_application_label(payment),
                "user": payment.created_by_name or "Histórico",
                "note": payment.note or "",
            })
        return {
            "rows": rows,
            "summary": {
                "payments_total": round(total_received, 2),
                "customers_with_balance": int(receivables["summary"].get("customers_with_balance") or 0),
                "overdue_total": round(float(receivables["summary"].get("overdue_total") or 0), 2),
                "due_soon_total": round(float(receivables["summary"].get("due_soon_total") or 0), 2),
            },
            "receivables": receivables,
        }

    def classify_expenses(self, start_date=None, end_date=None):
        query = Expense.query.filter(Expense.business_id == self.business_id)
        if start_date:
            query = query.filter(Expense.expense_date >= start_date)
        if end_date:
            query = query.filter(Expense.expense_date <= end_date)
        expenses = query.order_by(Expense.expense_date.desc(), Expense.id.desc()).all()

        debt_ids = sorted({expense.debt_id for expense in expenses if expense.debt_id})
        debts_by_id = {}
        if debt_ids:
            debts_by_id = {debt.id: debt for debt in Debt.query.filter(Debt.business_id == self.business_id, Debt.id.in_(debt_ids)).all()}

        totals = {
            "operational_expense": 0.0,
            "supplier_payment": 0.0,
            "operational_obligation_payment": 0.0,
            "financial_debt_payment": 0.0,
        }
        categories = {}
        rows = []

        for expense in expenses:
            amount = round(float(expense.amount or 0), 2)
            if amount <= 0:
                continue
            source_type = str(expense.source_type or "manual").strip().lower() or "manual"
            related_debt = debts_by_id.get(expense.debt_id)
            provider = "-"
            receipt = "-"
            status = "Registrado"
            flow_label = "Gasto operativo ejecutado"
            scope_label = "Operativo"

            if expense.supplier_payment and expense.supplier_payment.supplier:
                provider = expense.supplier_payment.supplier.name or "-"
                receipt = expense.supplier_payment.reference or f"Pago proveedor #{expense.supplier_payment_id}"
            elif expense.supplier_payable and expense.supplier_payable.supplier:
                provider = expense.supplier_payable.supplier.name or "-"
                receipt = expense.supplier_payable.raw_purchase.purchase_number if expense.supplier_payable.raw_purchase else f"CxP proveedor #{expense.supplier_payable_id}"
            elif expense.raw_purchase and expense.raw_purchase.supplier:
                provider = expense.raw_purchase.supplier.name or "-"
                receipt = expense.raw_purchase.purchase_number or f"Compra #{expense.raw_purchase_id}"
            elif related_debt and related_debt.creditor_name:
                provider = related_debt.creditor_name
                receipt = related_debt.name or f"Deuda #{related_debt.id}"
            elif expense.description:
                receipt = expense.description

            if source_type == "supplier_payment":
                totals["supplier_payment"] += amount
                flow_label = "Pago a proveedor"
                status = "Pagado"
            elif source_type == "debt_payment":
                scope = self.get_debt_scope(related_debt.category if related_debt else None)
                if scope == "financial":
                    totals["financial_debt_payment"] += amount
                    flow_label = "Pago de deuda financiera"
                    scope_label = "Financiero"
                else:
                    totals["operational_obligation_payment"] += amount
                    flow_label = "Pago de obligación operativa"
                status = (related_debt.status or "partial").capitalize() if related_debt else "Pagado"
            else:
                totals["operational_expense"] += amount
                flow_label = "Compra pagada" if source_type == "purchase_payment" else ("Gasto recurrente ejecutado" if source_type == "recurring" else "Gasto operativo ejecutado")
                category_key = expense.category or "Sin categoría"
                bucket = categories.setdefault(category_key, {"total": 0.0, "count": 0})
                bucket["total"] += amount
                bucket["count"] += 1
                if source_type == "recurring":
                    status = "Ejecutado"

            rows.append({
                "date": expense.expense_date,
                "category": expense.category or "Sin categoría",
                "description": expense.description or flow_label,
                "method": expense.payment_method or "-",
                "amount": amount,
                "provider": provider,
                "receipt": receipt,
                "status": status,
                "flow_label": flow_label,
                "scope": scope_label,
                "source_type": source_type,
            })

        debts = Debt.query.filter_by(business_id=self.business_id).all()
        operational_debts = [debt for debt in debts if self.get_debt_scope(debt.category) == "operational"]
        payables_summary = self.build_debts_summary(operational_debts, today=end_date or date.today())
        pending_recurring_total = float(
            db.session.query(func.coalesce(func.sum(RecurringExpense.amount), 0))
            .filter(
                RecurringExpense.business_id == self.business_id,
                RecurringExpense.is_active.is_(True),
                RecurringExpense.payment_flow == "payable",
            )
            .scalar() or 0
        )
        return {
            "rows": rows,
            "totals": {key: round(value, 2) for key, value in totals.items()},
            "categories": categories,
            "summary": {
                "expenses_total": round(sum(float(item["amount"] or 0) for item in rows), 2),
                "operational_expense_total": round(float(totals["operational_expense"] or 0), 2),
                "scheduled_or_pending_total": round(float(payables_summary.get("total_debt") or 0) + float(pending_recurring_total or 0), 2),
            },
            "payables_summary": payables_summary,
        }

    def resolve_sale_initial_cash_amount(self, sale, linked_payments, linked_ledger_payments):
        sale_date = getattr(sale, "sale_date", None)
        if not sale_date:
            return 0.0
        has_payment_on_sale_date = any(
            payment.payment_date == sale_date and float(payment.amount or 0) > 0
            for payment in linked_payments
        )
        if has_payment_on_sale_date:
            return 0.0
        ledger_initial_total = round(sum(
            float(entry.amount or 0)
            for entry in linked_ledger_payments
            if entry.entry_date == sale_date and float(entry.amount or 0) > 0
        ), 2)
        if ledger_initial_total > 0:
            return ledger_initial_total
        collected_amount = round(float(sale.collected_amount or 0), 2)
        if collected_amount <= 0:
            return 0.0
        if not linked_payments and not linked_ledger_payments:
            return collected_amount
        if not sale.customer_id:
            return collected_amount
        if round(float(sale.balance or 0), 2) <= 0.01:
            return collected_amount
        return 0.0

    def build_sale_initial_cash_events(self, period_start, period_end):
        if not period_start or not period_end:
            return []
        sale_rows = (
            Sale.query.options(joinedload(Sale.customer), joinedload(Sale.treasury_account))
            .filter(
                Sale.business_id == self.business_id,
                Sale.sale_date >= period_start,
                Sale.sale_date <= period_end,
                Sale.collected_amount > 0,
            )
            .order_by(Sale.sale_date.desc(), Sale.id.desc())
            .all()
        )
        sale_ids = [sale.id for sale in sale_rows]
        payments_by_sale_id = {}
        ledger_by_sale_id = {}
        if sale_ids:
            linked_payments = Payment.query.filter(Payment.business_id == self.business_id, Payment.sale_id.in_(sale_ids)).all()
            for payment in linked_payments:
                payments_by_sale_id.setdefault(payment.sale_id, []).append(payment)
            linked_ledger_entries = LedgerEntry.query.filter(
                LedgerEntry.business_id == self.business_id,
                LedgerEntry.ref_type == "sale",
                LedgerEntry.entry_type == "payment",
                LedgerEntry.ref_id.in_(sale_ids),
            ).all()
            for entry in linked_ledger_entries:
                ledger_by_sale_id.setdefault(entry.ref_id, []).append(entry)
        events = []
        for sale in sale_rows:
            amount = self.resolve_sale_initial_cash_amount(
                sale,
                payments_by_sale_id.get(sale.id, []),
                ledger_by_sale_id.get(sale.id, []),
            )
            if amount <= 0:
                continue
            customer_name = sale.customer.name if sale.customer else "Cliente casual"
            source_label = "Venta pagada" if round(float(sale.balance or 0), 2) <= 0.01 else "Abono inicial de venta"
            events.append({
                "date": sale.sale_date,
                "description": sale.note or f"Venta #{sale.id} • {customer_name}",
                "amount": round(float(amount or 0), 2),
                "category": sale.payment_method or "venta",
                "source_label": source_label,
                "movement_type": "income",
                "scope": "Operativo",
                "account": sale.treasury_account.name if sale.treasury_account else "-",
            })
        return events

    def build_cash_dataset(self, start_date=None, end_date=None):
        if not end_date:
            end_date = date.today()
        payments_query = Payment.query.options(joinedload(Payment.customer), joinedload(Payment.treasury_account)).filter(Payment.business_id == self.business_id)
        expenses_breakdown = self.classify_expenses(start_date, end_date)

        if start_date:
            payments_query = payments_query.filter(Payment.payment_date >= start_date)
        if end_date:
            payments_query = payments_query.filter(Payment.payment_date <= end_date)

        payment_rows = payments_query.order_by(Payment.payment_date.desc(), Payment.id.desc()).all()
        sale_initial_events = self.build_sale_initial_cash_events(start_date, end_date) if start_date and end_date else []

        movements = []
        for payment in payment_rows:
            movements.append({
                "date": payment.payment_date,
                "description": payment.note or f"Cobro cliente {payment.customer.name if payment.customer else payment.customer_id}",
                "amount": round(float(payment.amount or 0), 2),
                "category": payment.method or "pago",
                "source_label": "Cobro ejecutado",
                "movement_type": "income",
                "scope": "Operativo",
                "account": payment.treasury_account.name if payment.treasury_account else "-",
            })
        movements.extend(sale_initial_events)
        for expense in expenses_breakdown["rows"]:
            movements.append({
                "date": expense["date"],
                "description": expense["description"],
                "amount": round(float(expense["amount"] or 0), 2),
                "category": expense["category"],
                "source_label": expense["flow_label"],
                "movement_type": "expense",
                "scope": expense["scope"],
                "account": expense["method"] or "-",
            })
        movements.sort(key=lambda item: ((item["date"] or date.min), item["description"]), reverse=True)

        receivables = self.build_receivables_snapshot(reference_date=end_date)
        debts = Debt.query.filter_by(business_id=self.business_id).all()
        debts_summary = self.build_debts_summary(debts, today=end_date)
        operational_debts = [debt for debt in debts if self.get_debt_scope(debt.category) == "operational"]
        financial_debts = [debt for debt in debts if self.get_debt_scope(debt.category) == "financial"]
        operational_summary = self.build_debts_summary(operational_debts, today=end_date)
        financial_summary = self.build_debts_summary(financial_debts, today=end_date)

        cash_in = round(sum(item["amount"] for item in movements if item["movement_type"] == "income"), 2)
        cash_out = round(sum(item["amount"] for item in movements if item["movement_type"] == "expense"), 2)
        return {
            "summary": {
                "cash_in": cash_in,
                "cash_out": cash_out,
                "cash_net": round(cash_in - cash_out, 2),
                "accounts_receivable": round(float(receivables["summary"].get("total_pending") or 0), 2),
                "receivables_overdue_total": round(float(receivables["summary"].get("overdue_total") or 0), 2),
                "receivables_due_soon_total": round(float(receivables["summary"].get("due_soon_total") or 0), 2),
                "accounts_payable": round(float(debts_summary.get("total_debt") or 0), 2),
                "operational_payables_total": round(float(operational_summary.get("total_debt") or 0), 2),
                "financial_debt_total": round(float(financial_summary.get("total_debt") or 0), 2),
                "operational_expense_total": round(float(expenses_breakdown["totals"].get("operational_expense") or 0), 2),
                "supplier_payments_total": round(float(expenses_breakdown["totals"].get("supplier_payment") or 0), 2),
                "operational_obligation_payments_total": round(float(expenses_breakdown["totals"].get("operational_obligation_payment") or 0), 2),
                "financial_debt_payments_total": round(float(expenses_breakdown["totals"].get("financial_debt_payment") or 0), 2),
            },
            "movements": movements,
            "receivables": receivables,
            "expenses": expenses_breakdown,
            "debts_summary": debts_summary,
            "operational_summary": operational_summary,
            "financial_summary": financial_summary,
        }


def export_sales_excel(business_id, start_date=None, end_date=None):
    service = ReportExportService(business_id)
    start, end = service.normalize_dates(start_date, end_date)
    build_sales_report(service, start, end)
    return service.save("VENTAS")


def export_expenses_excel(business_id, start_date=None, end_date=None):
    service = ReportExportService(business_id)
    start, end = service.normalize_dates(start_date, end_date)
    build_expenses_report(service, start, end)
    return service.save("GASTOS")


def export_payments_excel(business_id, start_date=None, end_date=None):
    service = ReportExportService(business_id)
    start, end = service.normalize_dates(start_date, end_date)
    build_payments_report(service, start, end)
    return service.save("COBROS")


def export_cash_excel(business_id, start_date=None, end_date=None):
    service = ReportExportService(business_id)
    start, end = service.normalize_dates(start_date, end_date)
    build_cash_report(service, start, end)
    return service.save("CAJA")


def export_profitability_excel(business_id, summary_data, products_data, sales_data, alerts_data, start_date=None, end_date=None, filters=None):
    service = ReportExportService(business_id)
    start, end = service.normalize_dates(start_date, end_date)
    build_profitability_payload_report(
        service,
        summary_data,
        products_data,
        sales_data,
        alerts_data,
        start,
        end,
        filters=filters,
    )
    return service.save("RENTABILIDAD")


def export_team_excel(business_id, summary_data, detail_data, start_date=None, end_date=None):
    service = ReportExportService(business_id)
    start, end = service.normalize_dates(start_date, end_date)
    build_team_report(service, summary_data, detail_data, start, end)
    return service.save("REPORTE_EQUIPO")


def export_combined_report(business_id, report_type, start_date=None, end_date=None):
    normalized_type = str(report_type or "").strip().lower() or "general_business"
    service = ReportExportService(business_id)
    start, end = service.normalize_dates(start_date, end_date)
    if normalized_type == "sales_full":
        build_sales_report(service, start, end)
        return service.save("VENTAS_DETALLE")
    if normalized_type == "payments_full":
        build_payments_report(service, start, end)
        return service.save("COBROS")
    if normalized_type in {"finance_full", "cashflow_full"}:
        build_cash_report(service, start, end)
        prefix = "FINANCIERO" if normalized_type == "finance_full" else "FLUJO_CAJA"
        return service.save(prefix)
    if normalized_type == "general_business":
        build_general_business_report(service, start, end)
        return service.save("GENERAL")
    if normalized_type == "customers_full":
        build_customers_report(service, start, end)
        return service.save("CLIENTES_CARTERA")
    if normalized_type == "products_full":
        build_inventory_report(service, start, end)
        return service.save("INVENTARIO")
    if normalized_type == "aged_receivables":
        build_aged_receivables_report(service, start, end)
        return service.save("CARTERA_EDADES")

    from backend.services.export import export_combined_report as legacy_export_combined_report

    return legacy_export_combined_report(business_id, report_type, start_date, end_date)
