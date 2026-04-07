from copy import copy
from dataclasses import dataclass
from typing import Optional

from openpyxl.styles import Alignment, Border, Font, PatternFill, Side


@dataclass(frozen=True)
class CellStyleDefinition:
    font: Optional[Font] = None
    fill: Optional[PatternFill] = None
    border: Optional[Border] = None
    alignment: Optional[Alignment] = None
    number_format: Optional[str] = None


class ExcelStylePalette:
    def __init__(self, currency_code: str = "COP"):
        normalized_currency = str(currency_code or "COP").strip().upper() or "COP"
        self.currency_code = normalized_currency

        self._white = "FFFFFF"
        self._text = "111827"
        self._muted = "6B7280"
        self._border = "D1D5DB"
        self._soft_border = "E5E7EB"
        self._title = "0F172A"
        self._brand = "1D4ED8"
        self._brand_soft = "DBEAFE"
        self._header_fill = "1F2937"
        self._header_text = "FFFFFF"
        self._kpi_fill = "F8FAFC"
        self._filter_fill = "F3F4F6"
        self._total_fill = "E0F2FE"
        self._empty_fill = "F9FAFB"
        self._success_fill = "DCFCE7"
        self._success_text = "166534"
        self._warning_fill = "FEF3C7"
        self._warning_text = "92400E"
        self._danger_fill = "FEE2E2"
        self._danger_text = "991B1B"
        self._info_fill = "E0F2FE"
        self._info_text = "075985"

        thin = Side(style="thin", color=self._border)
        soft_thin = Side(style="thin", color=self._soft_border)
        medium = Side(style="medium", color=self._border)

        self.default_border = Border(left=soft_thin, right=soft_thin, top=soft_thin, bottom=soft_thin)
        self.header_border = Border(left=thin, right=thin, top=thin, bottom=thin)
        self.section_border = Border(left=medium, right=medium, top=medium, bottom=medium)

        self.number_formats = {
            "currency": f'"{self.currency_code}" #,##0.00',
            "integer": "0",
            "decimal": "0.00",
            "percentage": "0.00%",
            "date": "yyyy-mm-dd",
            "datetime": "yyyy-mm-dd hh:mm",
        }

        self.title = CellStyleDefinition(
            font=Font(name="Calibri", size=16, bold=True, color=self._title),
            alignment=Alignment(horizontal="left", vertical="center"),
        )
        self.subtitle = CellStyleDefinition(
            font=Font(name="Calibri", size=10, italic=True, color=self._muted),
            alignment=Alignment(horizontal="left", vertical="center"),
        )
        self.section_title = CellStyleDefinition(
            font=Font(name="Calibri", size=11, bold=True, color=self._brand),
            fill=PatternFill(fill_type="solid", start_color=self._brand_soft, end_color=self._brand_soft),
            border=self.section_border,
            alignment=Alignment(horizontal="left", vertical="center"),
        )
        self.table_header = CellStyleDefinition(
            font=Font(name="Calibri", size=10, bold=True, color=self._header_text),
            fill=PatternFill(fill_type="solid", start_color=self._header_fill, end_color=self._header_fill),
            border=self.header_border,
            alignment=Alignment(horizontal="center", vertical="center", wrap_text=True),
        )
        self.body = CellStyleDefinition(
            font=Font(name="Calibri", size=10, color=self._text),
            border=self.default_border,
            alignment=Alignment(horizontal="left", vertical="top", wrap_text=True),
        )
        self.body_center = CellStyleDefinition(
            font=Font(name="Calibri", size=10, color=self._text),
            border=self.default_border,
            alignment=Alignment(horizontal="center", vertical="top", wrap_text=True),
        )
        self.kpi_label = CellStyleDefinition(
            font=Font(name="Calibri", size=10, bold=True, color=self._muted),
            fill=PatternFill(fill_type="solid", start_color=self._kpi_fill, end_color=self._kpi_fill),
            border=self.header_border,
            alignment=Alignment(horizontal="left", vertical="center", wrap_text=True),
        )
        self.kpi_value = CellStyleDefinition(
            font=Font(name="Calibri", size=12, bold=True, color=self._title),
            fill=PatternFill(fill_type="solid", start_color=self._kpi_fill, end_color=self._kpi_fill),
            border=self.header_border,
            alignment=Alignment(horizontal="right", vertical="center"),
        )
        self.filter_label = CellStyleDefinition(
            font=Font(name="Calibri", size=10, bold=True, color=self._text),
            fill=PatternFill(fill_type="solid", start_color=self._filter_fill, end_color=self._filter_fill),
            border=self.default_border,
            alignment=Alignment(horizontal="left", vertical="center"),
        )
        self.filter_value = CellStyleDefinition(
            font=Font(name="Calibri", size=10, color=self._text),
            fill=PatternFill(fill_type="solid", start_color=self._empty_fill, end_color=self._empty_fill),
            border=self.default_border,
            alignment=Alignment(horizontal="left", vertical="center", wrap_text=True),
        )
        self.total_label = CellStyleDefinition(
            font=Font(name="Calibri", size=10, bold=True, color=self._title),
            fill=PatternFill(fill_type="solid", start_color=self._total_fill, end_color=self._total_fill),
            border=self.header_border,
            alignment=Alignment(horizontal="left", vertical="center"),
        )
        self.total_value = CellStyleDefinition(
            font=Font(name="Calibri", size=10, bold=True, color=self._title),
            fill=PatternFill(fill_type="solid", start_color=self._total_fill, end_color=self._total_fill),
            border=self.header_border,
            alignment=Alignment(horizontal="right", vertical="center"),
        )
        self.empty_state = CellStyleDefinition(
            font=Font(name="Calibri", size=10, italic=True, color=self._muted),
            fill=PatternFill(fill_type="solid", start_color=self._empty_fill, end_color=self._empty_fill),
            border=self.default_border,
            alignment=Alignment(horizontal="center", vertical="center", wrap_text=True),
        )
        self.footer = CellStyleDefinition(
            font=Font(name="Calibri", size=9, italic=True, color=self._muted),
            alignment=Alignment(horizontal="left", vertical="center"),
        )
        self.status_styles = {
            "success": CellStyleDefinition(
                font=Font(name="Calibri", size=10, bold=True, color=self._success_text),
                fill=PatternFill(fill_type="solid", start_color=self._success_fill, end_color=self._success_fill),
                border=self.default_border,
                alignment=Alignment(horizontal="center", vertical="center"),
            ),
            "warning": CellStyleDefinition(
                font=Font(name="Calibri", size=10, bold=True, color=self._warning_text),
                fill=PatternFill(fill_type="solid", start_color=self._warning_fill, end_color=self._warning_fill),
                border=self.default_border,
                alignment=Alignment(horizontal="center", vertical="center"),
            ),
            "danger": CellStyleDefinition(
                font=Font(name="Calibri", size=10, bold=True, color=self._danger_text),
                fill=PatternFill(fill_type="solid", start_color=self._danger_fill, end_color=self._danger_fill),
                border=self.default_border,
                alignment=Alignment(horizontal="center", vertical="center"),
            ),
            "info": CellStyleDefinition(
                font=Font(name="Calibri", size=10, bold=True, color=self._info_text),
                fill=PatternFill(fill_type="solid", start_color=self._info_fill, end_color=self._info_fill),
                border=self.default_border,
                alignment=Alignment(horizontal="center", vertical="center"),
            ),
        }

    def apply(self, cell, style: CellStyleDefinition, number_format: Optional[str] = None):
        if style.font is not None:
            cell.font = copy(style.font)
        if style.fill is not None:
            cell.fill = copy(style.fill)
        if style.border is not None:
            cell.border = copy(style.border)
        if style.alignment is not None:
            cell.alignment = copy(style.alignment)
        if number_format or style.number_format:
            cell.number_format = number_format or style.number_format or "General"

    def apply_named(self, cell, style_name: str, number_format: Optional[str] = None):
        style = getattr(self, style_name)
        self.apply(cell, style, number_format=number_format)

    def apply_status(self, cell, tone: str):
        style = self.status_styles.get(tone) or self.status_styles["info"]
        self.apply(cell, style)
