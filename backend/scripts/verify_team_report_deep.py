
import sys
import os
import io
import jwt
import random
from datetime import datetime, timedelta
import openpyxl

# Add the parent directory to the path
sys.path.append(os.path.abspath(os.path.join(os.path.dirname(__file__), '../..')))

from backend.main import app
from backend.database import db
from backend.models import User, Business, Role, TeamMember, Sale, Expense, Payment, Customer, ProductMovement, Product, Reminder, Permission, RolePermission
from backend.auth import create_token

def run_deep_verification():
    print("🚀 Starting FINAL REGRESSION: Team Report Verification...")
    
    with app.app_context():
        # --- 1. SETUP DATA ---
        print("\n--- 1. Setup Test Data (Deep) ---")
        
        # Create unique suffix to avoid collisions if run multiple times
        suffix = str(random.randint(1000, 9999))
        
        # Owner
        owner = User.query.filter_by(email=f"owner_{suffix}@test.com").first()
        if not owner:
            owner = User(email=f"owner_{suffix}@test.com", name="Owner Deep", password_hash="hash", plan="pro")
            db.session.add(owner)
            
        # Admin User
        admin_user = User.query.filter_by(email=f"admin_{suffix}@test.com").first()
        if not admin_user:
            admin_user = User(email=f"admin_{suffix}@test.com", name="Admin Deep", password_hash="hash")
            db.session.add(admin_user)

        # Finance User
        finance_user = User.query.filter_by(email=f"finance_{suffix}@test.com").first()
        if not finance_user:
            finance_user = User(email=f"finance_{suffix}@test.com", name="Finance Deep", password_hash="hash")
            db.session.add(finance_user)
            
        # Seller User (New)
        seller_user = User.query.filter_by(email=f"seller_{suffix}@test.com").first()
        if not seller_user:
            seller_user = User(email=f"seller_{suffix}@test.com", name="Seller Deep", password_hash="hash")
            db.session.add(seller_user)
            
        db.session.commit()
            
        # Business
        business = Business(user_id=owner.id, name=f"Deep Report Business {suffix}")
        db.session.add(business)
        db.session.commit()
        
        # --- ROLES & PERMISSIONS ---
        # Ensure Permission exists
        perm = Permission.query.filter_by(name="analytics.view_team").first()
        if not perm:
            perm = Permission(name="analytics.view_team", description="View Team Report", category="analytics")
            db.session.add(perm)
            db.session.commit()
            
        # 1. ADMIN Role (With Permission)
        role_admin = Role(name=f"ADMIN_{suffix}", business_id=business.id)
        db.session.add(role_admin)
        db.session.commit()
        
        # Assign permission to Admin
        rp = RolePermission(role_id=role_admin.id, permission_id=perm.id)
        db.session.add(rp)
        
        # 2. FINANCE Role (No Permission)
        role_finance = Role(name=f"FINANCE_{suffix}", business_id=business.id)
        db.session.add(role_finance)
        
        # 3. SELLER Role (No Permission)
        role_seller = Role(name=f"SELLER_{suffix}", business_id=business.id)
        db.session.add(role_seller)
        
        db.session.commit()
            
        # Team Members
        tm_admin = TeamMember(user_id=admin_user.id, business_id=business.id, role_id=role_admin.id, status="active")
        tm_finance = TeamMember(user_id=finance_user.id, business_id=business.id, role_id=role_finance.id, status="active")
        tm_seller = TeamMember(user_id=seller_user.id, business_id=business.id, role_id=role_seller.id, status="active")
        
        db.session.add_all([tm_admin, tm_finance, tm_seller])
        db.session.commit()

        # --- DATA GENERATION (Minimal for Export Check) ---
        # Create Customer
        customer = Customer(business_id=business.id, name="Paying Customer", created_by_user_id=owner.id, created_by_name=owner.name, created_by_role="Owner")
        db.session.add(customer)
        db.session.commit()
        
        # Payment
        payment = Payment(
            business_id=business.id,
            customer_id=customer.id,
            amount=200,
            payment_date=datetime.now().date(),
            method="cash",
            note="Deep Test Payment",
            created_by_user_id=owner.id,
            created_by_name=owner.name,
            created_by_role="Owner"
        )
        db.session.add(payment)
        db.session.commit()

        # Tokens
        owner_token = create_token(owner.id)
        admin_token = create_token(admin_user.id)
        finance_token = create_token(finance_user.id)
        seller_token = create_token(seller_user.id)
        
        client = app.test_client()

        print("\n--- 2. Validating Permissions ---")
        
        # 1. Owner Access
        res = client.get(f"/api/businesses/{business.id}/analytics/team", headers={"Authorization": f"Bearer {owner_token}"})
        print(f"🔹 OWNER Access: {res.status_code} (Expected 200)")
        if res.status_code != 200: print(f"❌ FAIL: {res.get_data(as_text=True)}")

        # 2. ADMIN Access (Has Permission)
        res = client.get(f"/api/businesses/{business.id}/analytics/team", headers={"Authorization": f"Bearer {admin_token}"})
        print(f"🔹 ADMIN Access (With Perm): {res.status_code} (Expected 200)")
        if res.status_code != 200: print(f"❌ FAIL: {res.get_data(as_text=True)}")

        # 3. FINANCE Access (No Permission)
        res = client.get(f"/api/businesses/{business.id}/analytics/team", headers={"Authorization": f"Bearer {finance_token}"})
        print(f"🔹 FINANCE Access (No Perm): {res.status_code} (Expected 403)")
        if res.status_code != 403: print(f"❌ FAIL: {res.status_code}")

        # 4. SELLER Access (No Permission)
        res = client.get(f"/api/businesses/{business.id}/analytics/team", headers={"Authorization": f"Bearer {seller_token}"})
        print(f"🔹 SELLER Access (No Perm): {res.status_code} (Expected 403)")
        if res.status_code != 403: print(f"❌ FAIL: {res.status_code}")

        print("\n--- 3. Validating Export ---")
        # 5. Export (Owner)
        res = client.get(f"/api/businesses/{business.id}/export/team", headers={"Authorization": f"Bearer {owner_token}"})
        print(f"🔹 EXPORT Access (Owner): {res.status_code} (Expected 200)")
        
        if res.status_code == 200:
            xls_content = io.BytesIO(res.data)
            wb = openpyxl.load_workbook(xls_content)
            if "Detalle Actividad" in wb.sheetnames:
                ws = wb["Detalle Actividad"]
                if ws.max_row > 1:
                    print("✅ Export Valid: Sheet 'Detalle Actividad' exists and has data.")
                else:
                    print("❌ Export Invalid: Sheet is empty.")
            else:
                print("❌ Export Invalid: Missing sheet.")
        else:
            print(f"❌ Export Failed: {res.status_code}")

if __name__ == "__main__":
    run_deep_verification()
